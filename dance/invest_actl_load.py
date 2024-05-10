#! /usr/bin/env python
'''
Update the table 'tbl_invest_actl'
'''
import argparse
import os
import sys

import pandas as pd
from dateutil.parser import parse
from openpyxl import load_workbook

from dance.other_actls import add_yyear_col
from dance.util.books import col_attrs_for_sheet, fresh_sheet, set_col_attrs
from dance.util.files import read_config, tsv_to_df
from dance.util.logs import get_logger
from dance.util.tables import (columns_for_table, conform_table,
                               df_for_table_name, write_table)
from dance.util.xl_formulas import dyno_fields

config=read_config()
logger=get_logger(__file__)

def read_passthru_accts(data_info):
  '''Grab the balances of the pass through accounts from the tsv file indicated in the argument
  
  argument: dictionary with 'path' to locate the file
  returns dataframe with the passthrough account names transformed to the actual account names and and their balances'''

  df=tsv_to_df(data_info['path'],skiprows=3,string_fields=['Account','Notes'])
  df=df.dropna(how='any')
  #strip off the indents
  df=pd.DataFrame([df['Account'].str.strip(),df['Current Balance']]).T # avoids warning A value is trying to be set on a copy of a slice from a DataFrame
  df=df.loc[df.Account.str.startswith('Pass-')]
  df['Account']=df.Account.replace('Pass-','',regex=True)# conform the passthru account names to the main account names
  df['Account']=df.Account.replace('-',' - ',regex=True)

  return df

def read_and_prepare_invest_actl(workbook,data_info,table_map=None):
  '''  Read investment actual data from files into a dataframe

  args:
    workbook: name of the workbook (to get the accounts and fees data from)
    data_info: dict that has a value for path used to locate the input file
    table_map: the dict that maps tables to worksheets. Required during initialization as it is not yet stored in file

  returns: dataframe with 6 rows per investment (add/wdraw, rlz int/gn, unrlz gn/ls, Income, Gains, Fees)
    The add/wdraw value is principal transfered, so interest payments have been removed if they are not reinvested.
    The unrealized gain value does not have the fees removed

  raises:
    FileNotFoundError: if the input file does not exist.
  '''

  logger.info('Starting investment actual load')

  xfer_file=data_info['path']
  perf_files_dir=data_info['file_sets']['performance']
  bals_files_dir=data_info['file_sets']['balances']

  # get the account list
  accounts=df_for_table_name(table_name='tbl_accounts',workbook=workbook,table_map=table_map)
  accounts=accounts[accounts.Type == 'I']

  # get the investment income in order to handle interest received from loans we hold
  invest_iande=df_for_table_name(table_name='tbl_invest_iande_values',workbook=workbook,table_map=table_map)

  # get the output of the Moneydance investment transfers report
  try:
    df=tsv_to_df (xfer_file,skiprows=3)
    logger.info('loaded dataframe from {}'.format(xfer_file))
  except FileNotFoundError as e:
    raise f'file not found {xfer_file}' from e
  df.dropna(axis=0,how='any',inplace=True)
  df=df.loc[~df.Account.str.contains('TRANSFERS')]
  df=df.groupby('Account').sum() # add the ins and outs
  df=df.mul(-1)# fix the sign
  df.index=df.index.str.strip() # remove leading spaces
  df=df.loc[:,df.columns !='Total'] # drop the total column
  cols=df.columns.to_list()
  map=dict(zip(cols,['Y'+a for a in cols]))
  df.rename(columns=map,inplace=True)
  y_columns=df.columns

  #adjust for loan interest received
  lir_data=[]
  lir_index=[]
  for account,row in accounts.loc[accounts['Reinv Rate']<1].iterrows():
    sel=(invest_iande.Account==account) & (invest_iande.Category.str.contains('Int:'))
    interest=invest_iande.loc[sel,y_columns].sum(axis=0).astype(float)
    adj= round(interest * (1-row['Reinv Rate']),2)
    adjusted= df.loc[account] + adj
    df.loc[account]=adjusted
    lir_data+=[adj.to_list()]
    lir_index+=[account]
  lir_df=pd.DataFrame(data=lir_data,index=lir_index,columns=y_columns)
    
  # join the accounts and transfers and keep only the data columns using the accounts list as the master
  xfers=accounts.join(df)[y_columns]
  xfers.fillna(0, inplace=True) # replace the NaN's (where no transfers occurred) with zeros

  # get the performance reports' data
  files=os.listdir(perf_files_dir)
  files.sort()
  file_cnt=0
  string_fields='Account,Security,Action,Unnamed: 8'.split(',')
  prior_pbals=None
  for file_name in files:
    if file_name=='.DS_Store': # MacOS finder leaves this around if you look at the folder.
      continue
    
    fn_year = file_name.split('.')[0].split('-')[-1] #grab the year from the file name
    perf_file=perf_files_dir+file_name
    logger.debug('Processing %s',fn_year)
    #read the internal date and check it.

    # now work with the performance data for a year
    df=tsv_to_df (perf_file,skiprows=3,string_fields=string_fields)
    # first check to see if the year is valid and its a full year
    orig=[df.columns[x] for x in [1,5]]
    open_close=[parse(x)for x in orig]
    ys=[x.year for x in open_close]
    assert ys[0]==ys[1], 'Report covers more than one year'
    assert 1+(open_close[1]-open_close[0]).days in [365,366], 'Not a full year'
    assert ys[0]==int(fn_year), 'File name year does not equal internal year'
    # change to 'Open' and 'End'
    map={orig[0]:'Open',orig[1]:'End'}
    df.rename(columns=map,inplace=True)

    #get the rows in order
    df.loc[:,'Security']=df['Security'].fillna(value='')
    df=df.loc[df.Security.str.startswith('Total:')] # throw away the securities, just keep the accounts
    df=df.loc[df.Security.str.strip() != 'Total:'] # remove grand total
    df.loc[:,'Security']=df['Security'].str.replace('Total:','').str.strip() # just the account names

    # prepare for joins with common index
    df.set_index('Security',inplace=True)
    found=[x in accounts.index.to_list() for x in df.index.to_list()]
    if not all(found):

      print('Not all accounts for %s are in master. Missing:'%fn_year)
      for a in [x for x in df.index.to_list() if x not in accounts.index.to_list()]:
        print('  %s'%a)
      assert all(found)
    df=df.drop(['Buys','Sales','Return %','Annual % (ROI)'],axis=1)# drop some columns we don't care about
    # all master accts, and all performance columns
    df=accounts.join(df)[df.columns]
    df.fillna(value=0,inplace=True) # zero out the perf values for accts in master not perf

    # append the transfers
    df=df.join(xfers['Y'+fn_year])
    df.rename(columns={'Y'+fn_year:'Transfers'},inplace=True)

    # get the fees and append as a column
    iiw=pd.read_excel(workbook,sheet_name='invest_iande_work',skiprows=1)
    iiw=iiw.loc[iiw.isnull().all(axis=1).cumsum()==0] # drop the ratios
    iiw=iiw.loc[iiw.Category.isin(['Investing:Account Fees'])]#Not including 'Investing:Action Fees'
    iiw=iiw.reset_index().pivot_table(index='Account',values='Y'+fn_year,aggfunc='sum')# add both type of fees together
    df=df.join(iiw['Y'+fn_year],how='left').fillna(value=0)
    df.rename(columns={'Return Amount': 'Return_Amount','Y'+fn_year:'Fees'},inplace=True)

    # put the loan interest received in a column
    df=df.join(lir_df['Y'+fn_year],how='left').fillna(value=0)
    df.rename(columns={'Y'+fn_year:'LIR'},inplace=True)

    # get the pending re-investments from the Merrill IRA
    retro=['IRA - VEC - ML'] # the pending only makes sense for accounts where the outgoing $ comes back as securities - i.e. this Merrill account
    pbals=read_passthru_accts({'path':bals_files_dir+file_name}) 
    pbals=pbals.loc[pbals.Account.isin(retro)].copy()
    pbals.set_index('Account',inplace=True)# same index for the join
    pbals.rename({'Current Balance':'Pending'},axis=1,inplace=True)# convenient column name
    df=df.join(pbals,how='left').fillna(value=0) # df now has pending column

    if prior_pbals is not None:
      df=df.join(prior_pbals,how='left').fillna(value=0) # df now has prior column
    else:
      df['Prior']=0
    prior_pbals=pbals.copy() # for next year
    prior_pbals.rename({'Pending':'Prior'},axis=1,inplace=True)# convenient column name
    

    # compute the two gain fields
    df=df.assign(Realized=lambda x: x.Income + x.Gains) # compute the realized gains

    # Note: the gains includes the fees associated with each sale as part of the cost basis
    # And, is believed to include the same on the buy side.

    df=df.assign(Unrealized=lambda x: (x.Return_Amount- (df.Pending - df.Prior)) - x.Realized)# unrealized is a plug (does not include fees)

    df['Check']=(df.Open + df.Transfers + df.Realized + df.Unrealized +df.Fees - df.LIR).round(2)
    df['Delta']=df.End - df.Check
    valid= (df.End-df.Check).abs() < .001
    if ~valid.all():
      logger.error('We have a problem getting the end balances. File = {}'.format(file_name))
      print(df.loc[~valid])
      print('''
      Things to look at
      - Merrill (IRA) reinvestment rounding problems - off by 1 or 2 cents.
        - Export Transfers, Detailed report and run dance.util.match_retro.py to locate
      - Unit prices - not same precision 
        - At price history don't take the printed unit value - recalculate as ratio for 12/31/yy price on items that are off
      - Income items not marked as MiscInc 
        - such as interest being coded as xfr
        - remember to re-run/save performance report
      - Transfers that do not pass through a bank (or items in the transfers report used to generate the invest_x.tsv file)
      - The Transfers, Detailed report for just the account and year can be helpful.
      - Transfers to/from investment accounts should use xfr not buyxfr or sellxfr.
      ...''')
      yn='Q'
      while not yn in 'YN':
        yn = input("Continue? (N/y):").strip().upper()+'N'
        if yn=='N':
          sys.exit(-1)
      logger.info(f'Continuing, with known Investment balance error for {file_name}')
    else:
      logger.info('Investment balance checks OK for {}'.format(file_name))
    
    # remove the loan interest received column
    df.drop(columns='LIR',inplace=True)

    # use names in spreadsheet
    map={'Transfers':'Add/Wdraw','Realized':'Rlz Int/Gn','Unrealized':'Unrlz Gn/Ls'}
    df.rename(columns=map,inplace=True)

    # rework so the rows of 6 value types for each account
    rows=df[['Add/Wdraw','Rlz Int/Gn','Unrlz Gn/Ls','Income','Gains','Fees']].transpose().stack().reset_index()
    map=dict(zip(rows.columns.to_list(),['ValType','Account','Y'+fn_year]))
    rows.rename(columns=map,inplace=True)
    rows=rows.assign(Key=lambda x: x.ValType + x.Account)
    rows.set_index('Key',inplace=True)
    if file_cnt==0:
      table=rows
    else:
      table=table.join(rows[rows.columns.to_list()[-1]])
    file_cnt=file_cnt+1
  # put the data into the spreadsheet
  wb=load_workbook(filename =workbook)
  table.reset_index(inplace=True) # puts the key back into 1st column
  col_def=columns_for_table(wb,'invest_actl','tbl_invest_actl',config)
  table=conform_table(table,col_def['name'])

  return table

def get_cap_gains(data_info):
  '''Get cap gains'''
  txt_fields=['Account', 'Security', 'Action','Unnamed: 8']
  df=tsv_to_df (data_info['path'],skiprows=3,string_fields=txt_fields)
  df['Account'].fillna(method='ffill',inplace=True)
  df['Security'].fillna(method='ffill',inplace=True)
  df=df.loc[~pd.isnull(df.Action)]
  df=df.loc[~df['Action'].str.startswith('Total:')]
  if False:
    #create Duration column and populate Action column for duration rows
    durations=['Long Term','Short Term','Unrealized']
    for _,row in   df.iterrows():
      action=row['Action']
      if action.startswith('Sell') or action=='Unrealized':
        share_info=row['Unnamed: 8']
        share_info=parse_share_info(share_info)
        block=df.loc[(df.Account==row['Account']) & (df.Date==row['Date']) & (df.Action.isin(durations[:-1]))]
        pass
        for iy,row2 in block.iterrows():
          duration=row2['Action']
          assert duration in durations, 'Expecting Short Term, Long Term, or Unrealized for the action, got: ' + duration
          assert row2['Shares']>=0, 'eh? negative shares'
          if share_info:
            assert row2['Shares']<=share_info[duration],'Inconsistent information for shares/action'
          df.at[iy,'Duration']=duration
          real={'Sell':'Realized','SellXfr':'Realized','Unrealized':'Unrealized'}[action]
          df.at[iy,'Realized']=real # copy the main records action to the block record so its on the same row as the duration
          # the main action row will have a nan in the Realized column. 
          # and thus won't be included in the summary, since its accounted for on the block record with the duration
  del df['Unnamed: 8']
  df=add_yyear_col(df)
  #df1=df.loc[~pd.isnull(df.Duration)]
  df.loc[df.Action=='SellXfr','Action']='Sell'
  df.loc[df.Action=='Sell','Action']='Realized'
  df=df.loc[~df.Action.str.endswith('Term')]
  summary=df.pivot_table(index=['Account'],aggfunc='sum',values='Gains',columns=['Action','Year']) 
  pass
  # Notes:
  #  return amount column on Investment Performance seems to include unrealized, realized and income
  #  average cost method seems to return non sensical results for unrealized
  #   with possible exception when entire lot is sold - like in BKG - JNT - ML
  


  return summary

def parse_share_info(share_info):
  result=None
  if not pd.isna(share_info):
    share_info=share_info.split(':')[-1].strip()
    share_info=share_info.replace(' = ','=')
    share_info=share_info.replace('  ',' ')
    result={}
    for trm in share_info.split(' '):
      dur,shrs=trm.split('=')
      dur={'LT':'Long Term','ST':'Short Term'}[dur]
      shrs=float(shrs.replace(',',''))
      result[dur]=shrs
  return result

if __name__ == '__main__':
  default_wb=config['workbook']
  parser = argparse.ArgumentParser(description ='Copies data from input files into tab "invest_actl". ')
  parser.add_argument('--workbook','-w',default=default_wb,help=f'Target workbook. Default: {default_wb}')
  parser.add_argument('--input','-i',default= 'data/invest-x.tsv',help='The path to the balances files')
  parser.add_argument('--balances','-b',default= 'data/acct-bals/',help='The path to the balances files')
  parser.add_argument('--performance','-f',default= 'data/invest-p/',help='The path to the performance reports')
 

  args=parser.parse_args()
  sheet='invest_actl'
  table='tbl_'+sheet
  data_info={'path': os.getcwd()+'/'+args.input,
    'file_sets':{'balances':args.balances,'performance':args.performance}}
  data=read_and_prepare_invest_actl(workbook=args.workbook,data_info=data_info)
  table_info=read_config()['sheets'][sheet]['tables'][0]
  data=dyno_fields(table_info,data) # get the taxable status
  wkb = load_workbook(filename = args.workbook, read_only=False)
  wkb=fresh_sheet(wkb,sheet)
  wkb= write_table(wkb,target_sheet=sheet,df=data,table_name=table)
  attrs=col_attrs_for_sheet(wkb,sheet,read_config())
  wkb=set_col_attrs(wkb,sheet,attrs)
  wkb.save(args.workbook)
