#!/usr/bin/env python

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
import openpyxl.utils.cell
from util.logs import get_logger

def tsv_to_df(filename,sep='\t',skiprows=0):
  '''grab the data from a Moneydance report and return a dataFrame'''
  logger=get_logger(__file__)
  try:
    df=pd.read_csv(filename,sep=sep,skiprows=skiprows,dtype='str') #,index_col=0)
    # keep as string because there are some clean ups needed
  except FileNotFoundError:
    logger.error(f'No file "{filename}". Quitting.')
    quit()
  cols=df.columns
  for col in cols[1:]:
    if col != 'Date' and col != 'Notes':
      df.loc[:,col]=df[col].fillna(value='0.00')
      df.loc[:,col]=df[col].str.replace(r'\$','',regex=True)
      df.loc[:,col]=df[col].str.replace('--','0.00') # occurs in performance report
      df.loc[:,col]=df[col].str.replace('None','0.00') # occurs in performance report
      df.loc[:,col]=df[col].str.replace(',','').astype(float)
    if col == 'Date':
      df[col] = pd.to_datetime(df[col])

  return df


def fresh_sheet(wb,sheet_name,tab_color='58BD2D'):
  ''' remove the named sheet from the work book and created a fresh one at the same location, 
  copying tab color.  If not already existing will create the new sheet at the end.
  returns the workbook
  '''
  logger=get_logger(__file__)
  ix=len(wb.sheetnames)
  if sheet_name in wb.sheetnames:
    ix= wb.sheetnames.index(sheet_name)
    ws = wb[sheet_name]
    tab_color=ws.sheet_properties.tabColor
    wb.remove(ws)
    logger.info('deleted worksheet {}'.format(sheet_name))
  ws=wb.create_sheet(sheet_name,ix)
  logger.info('created worksheet {}'.format(sheet_name))
  ws.sheet_properties.tabColor=tab_color
  return wb

def df_for_range(worksheet=None,range_ref=None):
  '''get a dataframe given a range reference on a worksheet'''
  
  from itertools import islice
  data=[]
  rb=openpyxl.utils.cell.range_boundaries(range_ref)
  for src_row in worksheet.iter_rows(min_col=rb[0],min_row=rb[1], max_col=rb[2], max_row=rb[3]):
    row=[]
    for cell in src_row:
      row.append(cell.value)
    data.append(row)
  
  cols = data[0][1:]
  idx = [r[0] for r in data[1:]]
  data = [r[1:] for r in data[1:]]
  df = pd.DataFrame(data, index=idx, columns=cols)
  return df

def get_val(table=None, line_key=None ,  col_name = None):
  return table.loc[line_key,col_name]

def put_val(table=None, line_key=None ,  col_name = None, value=None):
   table.loc[line_key,col_name]=value


def ws_for_table_name(table_map=None,table_name=None):
  '''return name of worksheet holding table given a data frame holding the table'''

  return table_map.loc[table_name,'Worksheet']

def df_for_table_name(table_name=None, data_only=False):
  '''Opens the file, and extracts a table as a dataFrame with the first column as the dataframe index'''
  logger=get_logger(__file__)
  source='data/fcast.xlsm'
  wb = load_workbook(filename = source, read_only=False, keep_vba=True, data_only=data_only)
  ws=wb['utility']
  ref=ws.tables['tbl_table_map'].ref
  table_map = df_for_range(worksheet=ws,range_ref=ref)
  ws_name =ws_for_table_name(table_map=table_map, table_name=table_name)
  ws=wb[ws_name]
  table=df_for_range(worksheet=ws,range_ref=ws.tables[table_name].ref)
  logger.info('Loaded worksheet {} from {}'.format(ws_name,source))
  logger.info('{} rows and {} columns'.format(table.shape[0],table.shape[1]))
  return table

def df_for_table_name_for_update(table_name=None):
  '''Opens the file, and extracts a table as a dataFrame with the first column as the dataframe index
  returns dataframe, worksheet, range_ref and workbook
  since its going to be updated don't allow data only '''
  logger=get_logger(__file__)
  source='data/fcast.xlsm'
  wb = load_workbook(filename = source, read_only=False, keep_vba=True)
  logger.info('Loaded workbook from {}'.format(source))
  ws=wb['utility']
  ref=ws.tables['tbl_table_map'].ref
  table_map = df_for_range(worksheet=ws,range_ref=ref)
  ws_name =ws_for_table_name(table_map=table_map, table_name=table_name)
  ws=wb[ws_name]
  table=df_for_range(worksheet=ws,range_ref=ws.tables[table_name].ref)
  logger.info('{} rows and {} columns in sheet {}'.format(table.shape[0],table.shape[1],ws))
  return table, ws,ws.tables[table_name].ref,wb

def main():
  '''modify this to test new functions'''

  df=tsv_to_df('data/2018 Account Balances.tsv',skiprows=3)
  print(df)

if __name__ == "__main__":
    """ execute for testing by running as a script"""

    main()    