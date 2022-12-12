#! /usr/bin/env python
'''For actual years, loads the bank, credit card and non-bank transfers into the 'transfers_actl' tab

This handles creating nested account keys, groups and totals as well as the raw data.

'''
import argparse
import pandas as pd
import numpy as np

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from dance.bank_actl_load import bank_cc_changes
from dance.util.books import fresh_sheet, col_attrs_for_sheet,set_col_attrs
from dance.util.files import  tsv_to_df,read_config
from dance.util.tables import get_f_fcast_year,write_table,columns_for_table
from dance.util.logs import get_logger

logger=get_logger(__file__)

def read_transfers_actl(data_info,target_file='data/fcast.xlsm',table_map=None):
  '''  Read data from files into a dataframe

  args:
    data_info: dict that has a value for path used to locate the input file
    target_file: needed to remove interest from the bank accounts
    table_map: in case the file is under construction and has not yet been written

  returns: dataframe with index = the account key

  raises:
    FileNotFoundError: if the input file does not exist.
  '''
  input_file=data_info['path']
  try:
    df=tsv_to_df (input_file,skiprows=3,nan_is_zero=False)
    logger.info('loaded dataframe from {}'.format(input_file))
  except FileNotFoundError as e:
    raise f'file not found {input_file}' from e

  df=df[~df['Account'].isnull()] #remove the blank rows
  df=df[~df['Account'].str.contains('TRANSFERS')] # remove the headings and totals

  # throw away total column
  df.rename(columns={'Total':'level'},inplace=True)

  # create the key field to allow for pivot
  # fill it first with a temporary value used to compute the heirarchy level
  df.insert(loc=0,column='key',value=df.Account.str.lstrip())
  #create a level indicator as a dataframe field, then as a list
  df['level']=((df.Account.str.len() - df.key.str.len())-4)/3
  level=[int(x) for x in df.level.tolist()]
  del df['level']

  keys=[]
  keyparts=df.Account.tolist()
  rows=list(range(0,len(level))) # rows and columns are origin 0, excel uses origin 1 plus it contains the headings so for rows its off by 2
  last_level=-1
  for rw in rows:
    lev=level[rw]
    if 0==lev:
      pathparts=[]
    else:
      if lev > last_level: pathparts.append(keyparts[rw-1].strip())
      if lev < last_level:
        pathparts=pathparts[:-1]
    a=pathparts.copy()
    a.append(keyparts[rw].strip())
    keys.append(':'.join(a))
    last_level=lev

  # put the keys into the df and make that the index
  df['key']=keys
  df.set_index('key',inplace=True)
  del df ['Account'] # its in the key

  # change the year columns to Y+year format
  for cn in df.columns.values.tolist():
    try:
      n=int(cn)
      nn = 'Y{}'.format(n)
      df.rename(columns={cn:nn},inplace=True)
    except ValueError:
      n=0

  # The parent accounts don't contain data.  Get a list of those
  parents=pd.unique(df.loc[df.isnull().all(axis=1)].index)

  # summary adds the inbound and outbound transfers together
  summary=pd.pivot_table(df,index=['key'],values=df,aggfunc=np.sum)
  summary=summary.mul(-1)

  # mark the parent rows, since the sum function turns their values into zeros
  summary.loc[summary.index.isin(parents),summary.columns]=np.nan

  # bring in the bank data
  bank_changes=bank_cc_changes(data_info=data_info,target_file=target_file,table_map=table_map)

  #combine the sets
  df =pd.concat([summary,bank_changes])
  df.sort_index(inplace=True)
  #df=df.reset_index().rename(columns={'key':'Account'})
  return df

def prepare_transfers_actl(workbook,df,f_fcast=None):
  '''prepare the dataframe for insertion into workbook.
  Handles category nesting as groups with subtotals.

  args:
    workbook: the name of the spreadsheet to load data into.
    df: the dataframe that has basic clean up already done
    f_fcast: Optional. The first forecast year as Ynnnn. If none, will read from the workbook file. Default None.

  returns:
    A dataframe and the groups (for folding)

  raises:
    FileNotFoundError: if workbook does not exist or the config file does not exist.
    IndexError: if columns expected and received from data source do not match
  '''
  # get the workbook from file
  try:
    wb = load_workbook(filename = workbook, read_only=False, keep_vba=True)
    logger.info('loaded workbook from {}'.format(workbook))
    config=read_config()
  except FileNotFoundError:
    raise FileNotFoundError(f'file not found {workbook}') from None

  # validate only one table
  target_sheet= 'transfers_actl'
  tables=config['sheets'][target_sheet]['tables']
  assert len(tables)==1,'not exactly one table defined'
  table_info=tables[0]

  # set defaults
  if f_fcast is None:
    f_fcast=get_f_fcast_year(wb,config) # get the first forecast year from the general state table or config as available
  logger.info ('First forecast year is: %s',f_fcast)
  key_values={'title_row':1,'start_col':1,'include_years':False}
  for k in key_values:   #take specified or default items
    if k in table_info:
      key_values[k]=table_info[k]

  # The parent accounts don't contain data.  Get a list of those
  parents=list(df.loc[df[df.columns[1:]].isnull().all(axis=1)].index)

  groups=[] # level, start, end
  #keep track of nesting of accounts
  heir=[]
  stot=[]
  rw=0
  xl_off=key_values['title_row']+1
  keys=df.index
  level=[x.count(':')for x in keys.tolist()]
  next_level=level[1:]+[0]
  gr_df=pd.DataFrame([],columns=df.columns) # a new dataframe with the groups and subtotals
  for ky in keys:
    # if the key or any of its parts an exact match to one of the parents then we defer writing it
    if any([x==ky for x in parents]): 
      heir.append(ky)# the key and the subtotal count
      stot.append(0)
    else:
      #determine how far back to look for the subtotal
      pending=len(heir)
      xl_row=1+rw+xl_off-pending

      #write the values from this row
      data_row=df.loc[[ky]]
      gr_df=pd.concat([gr_df,data_row])
      #process subtotals if there are any
      if pending>0:
        stot[-1]=stot[-1]+1 # include the row we just wrote in the subtotal

        # if the next row is at a lower level, then insert a total row for each time it pops
        pop_count=level[rw]-next_level[rw]
        adj_row=0
        while pop_count>0:
          nr_items_to_st=stot[-1]

          # subtotal row location adds one for each pop
          adj_row=adj_row+1
          stot_row=xl_row+adj_row

          # locate the start of the items to subtotal
          xl_start_row= stot_row - nr_items_to_st

          # insert the label and the formulas
          formulas=[]
          for cl in range(gr_df.shape[1]):
            let=get_column_letter(cl+key_values['start_col']+1)# data has the years, account is in index now but it becomes the A column.
            formula='=subtotal(9,{}{}:{}{})'.format(let,xl_start_row,let,stot_row-1)
            formulas.append(formula)
          st_df=pd.DataFrame([formulas],columns=gr_df.columns,index=[heir[-1] + ' - TOTAL'])
          gr_df=pd.concat([gr_df,st_df])
          # capture the grouping info
          groups.append([level[rw]-(adj_row-1),xl_start_row,stot_row-1])

          # pop the items
          heir.pop()
          stot.pop()
          pop_count=pop_count-1

          # include the count in the next higher level
          if len(stot)>0:
            stot[-1]=stot[-1]+nr_items_to_st+1
    rw=rw+1

  # sort the row groups definitions: highest level to lowest level
  def getlev (e):
    return e[0]

  # use a histogram to determine where each group's foot print overlaps with the span of other groups
  hist=[0]*max([rw]+[x[2]for x in groups]) # handle case when last row is still indented
  groups.sort(key=getlev)
  for grp in groups:
    foot=list(range(grp[1]-1,grp[2]))
    mx=1+max([hist[x]for x in foot])
    grp[0]=mx
    for x in foot:
      hist[x]=hist[x]+1

  groups.sort(key=getlev)
  gr_df=gr_df.reset_index()
  gr_df.rename(columns={'index':'Account'},inplace=True)

  # validate columns (after we move Account from index to columns)
  tab_tgt=table_info['name']
  cols_df=columns_for_table(wb,target_sheet,tab_tgt,config)
  expected=set(cols_df.name)
  data_has=set(gr_df.columns)
  if data_has != expected:
    msg='Columns expected does not match data source for {}'.format(tab_tgt)
    logger.error(msg)
    logger.error('Extra expected columns: {}, Extra received columns: {}'.format(expected-data_has,data_has-expected))
    raise IndexError(msg)

  return gr_df,groups

if __name__ == '__main__':
  parser = argparse.ArgumentParser(description ='Copies data from input file into tab "transfers_actl". ')
  parser.add_argument('--workbook','-w',default='data/fcast.xlsm',help='Target workbook')
  parser.add_argument('--path','-p',default= 'data/transfers.tsv',help='The path and name of the input file')
  parser.add_argument('--ffy', '-y',help='first forecast year. Must be provided if workbook does not have value. Default None.')
  args=parser.parse_args()
  df=read_transfers_actl(data_info={'path':args.path},target_file=args.workbook)
  transfers_actl,fold_groups=prepare_transfers_actl(args.workbook,df)
  sheet='transfers_actl'
  table='tbl_'+sheet
  wkb = load_workbook(filename = args.workbook, read_only=False, keep_vba=True)
  wkb=fresh_sheet(wkb,sheet)
  wkb= write_table(wkb,target_sheet=sheet,df=transfers_actl,table_name=table,groups=fold_groups)
  attrs=col_attrs_for_sheet(wkb,sheet,read_config())
  wkb=set_col_attrs(wkb,sheet,attrs)
  wkb.save(args.workbook)
