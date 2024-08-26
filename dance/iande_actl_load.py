#! /usr/bin/env python
'''
Copies data from "data/iande.tsv" into tab "iande" after doing some checks.
'''
import argparse
import os

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from dance.util.books import col_attrs_for_sheet, fresh_sheet, set_col_attrs
from dance.util.files import read_config, tsv_to_df
from dance.util.logs import get_logger
from dance.util.tables import (columns_for_table, df_for_table_name,
                               get_f_fcast_year,  write_table)
from dance.util.xl_formulas import actual_formulas, forecast_formulas, table_ref
from dance.util.row_tree import hier_insert,folding_groups,is_leaf,nest_by_cat,subtotal_formulas,set_level

config=read_config() 
logger=get_logger(__file__)


def indent_other(str):
  '''indent -other rows to fix an anomoly in the md report
  '''
  if -1==str.find('-Other'):
    return str
  else:
    return '   '+str

def read_iande_actl(data_info):
  '''  Read data from file into a dataframe

  args:
    data_info: dict that has a value for path used to locate the input file

  returns: dataframe with index as 0...n

  raises:
    FileNotFoundError: if the input file does not exist.
  '''
  input_file=data_info['path']

  # get the input data from file
  try:
    df=tsv_to_df (input_file,skiprows=3)
    logger.info('loaded dataframe from {}'.format(input_file))
  except FileNotFoundError as e:
    raise f'file not found {input_file}' from e

  # some data clean ups
  df.fillna(0, inplace=True) # replace the NaN's with zeros
  df.query('Account!=0',inplace=True) # remove blank rows
  df.Account=df.Account.apply(indent_other)
  if 'Total' in df.columns:
    del df['Total']
  df.reset_index(drop=True,inplace=True)
  return df

def y_year(df):
  '''change the year columns to Y+year format'''
  for cn in df.columns.values.tolist():
    if cn.isnumeric():
      n=int(cn)
      nn = 'Y{}'.format(n)
      df.rename(columns={cn:nn},inplace=True) 
  return df 

def current_inclusive(current: pd.DataFrame, wb:Workbook ,table_map:dict)-> pd.DataFrame:
  '''For current iande, expand the row set to include what is already in the workbook for iande.
  This allows reforecasting a row that does not yet exist in the year but has existed in past years.
  
  We use a left join, so this will exclude any rows in current that don't exist in iande.
  This is desired since the ytd function will try to post values from current to iande.

  The resulting dataframe has two non-null fields: Key and level.

  Provide the ytd dataframe as current, and the workbook
  Returns revised current dataframe with new rows.
  '''

  ref=df_for_table_name('tbl_iande',wb,table_map=table_map)[[]] # just the index
  ref.index.name='Key'
  ref=set_level(ref) 
  current.set_index('Key',inplace=True)
  del current['level'] # since its a left join, we can just use the ones from iande
  current=ref.merge(current,on='Key',how='left')
  current.reset_index(inplace=True)
  return current

def prepare_iande_actl(workbook,target_sheet,df,force=False,f_fcast=None,title_row=1,verbose=False,
                       table_map:dict = None):
  '''Prepare the dataframe for insertion into workbook. 

    Handles category nesting as groups with subtotals.

    args:
      workbook: the name of the spreadsheet to load data into.
      target_sheet: the name of the tab to update. Either 'iande' or 'current'
      df: the dataframe that has basic clean up already done
      force: Optional. True to override warning. Default False
      f_fcast: Optional. The first forecast year as Ynnnn. If none, will read from the workbook file. Default None.
      title_row: Optional. The row number in Excel for the title row (needed to compute subtotals). Default 1.
      verbose: True to report out all the lines that are in actl but not in iande. default False.
      table_map is required when building the whole book (to support the current tab, using iande).

    returns:
      A dataframe and the groups (for folding)

    raises:
      FileNotFoundError: if workbook does not exist or the config file does not exist.
      ValueError: if tab_target is not iande or current
      IndexError: if columns expected and received from data source do not match
    '''

  if title_row!=1:
    raise NotImplementedError('Title row must be 1 for {} for now.'.format(target_sheet))
  valid_sheets=['iande','current']
  heading_row=1+title_row
  if target_sheet not in valid_sheets:
    raise ValueError('tab_target must be one of: %s'%', '.join(valid_sheets))
  logger.debug('Preparing data for {}'.format(target_sheet))
  # get the workbook from file
  try:
    wb = load_workbook(filename = workbook, read_only=False)
    logger.debug('loaded workbook from {}'.format(workbook))
    config=read_config()
  except FileNotFoundError:
    raise FileNotFoundError(f'file not found {workbook}') from None
  if f_fcast is None:
    f_fcast='Y%d' % get_f_fcast_year(wb,config) # get the first forecast year from the general state table or config as available
  logger.debug ('First forecast year is: %s',f_fcast)

  tables=config['sheets'][target_sheet]['tables']
  assert len(tables)==1,'not exactly one table defined'
  tab_tgt=tables[0]['name']

  df=nest_by_cat(df) # creates key and level fields
  df=hier_insert(df,tables[0]) # insert any specified lines into hierarchy

  if target_sheet=='current':
    df=current_inclusive(df,wb,table_map=table_map)

  df=is_leaf(df)# mark rows that are leaves of the category tree
  groups=folding_groups(df)
  del df['level'] # clear out temp field  

  expected_cols=set(columns_for_table(wb,target_sheet,tab_tgt,config).name)

  match target_sheet:
    case 'iande':
      del df['is_leaf'] # clear out temp field
      
      for y in range(int(f_fcast[1:]),config['end_year']+1): # add columns for forecast years
        df['Y{}'.format(y)]=None

    case 'current':
      # verify that year is the first forecast year
      as_of=df.columns[-1]
      as_of_datetime=pd.to_datetime(as_of.split(' - ')[-1])
      assert as_of_datetime.year==int(f_fcast[1:]),'YTD file is not for first forecast year'

      # rename the ytd column to correct format
      yymmdd='%4d%02d%02d'%(as_of_datetime.year,as_of_datetime.month,as_of_datetime.day)
      df.rename(columns={as_of:yymmdd},inplace=True)

      expected_cols= (expected_cols-set(['YTD']))|set(['Y'+yymmdd]) 
      df[['Factor','Add','Year']]=None

      # set the formula for the total year
      formula='=IF(AND(ISBLANK([@Factor]),ISBLANK(@Add)),"",([@Y%s]*[@Factor])+[@Add])'%yymmdd
      df['Year']=table_ref(formula)          
      pass
    
  df=y_year(df)  
  df=subtotal_formulas(df,groups)

  # support for iande etc.
  keys=df.Key.tolist()
  grand_total='TOTAL INCOME - EXPENSES' 
  if grand_total in keys:
    net_ix=keys.index(grand_total) # find the net line (its should be the last line)
    inc_ix=keys.index('Income - TOTAL')+heading_row+1 # offset for excel
    exp_ix=keys.index('Expenses - TOTAL')+heading_row+1
    for cx,cl in enumerate(df.columns):
      if str(cl).startswith('Y'):
        let=get_column_letter(cx+1)
        formula='={}{}-{}{}'.format(let,inc_ix,let,exp_ix)
        df.loc[net_ix,[cl]]=formula

  data_has_cols=set(df.columns)
  if data_has_cols != expected_cols:
    msg='Columns expect does not match data source for {}'.format(tab_tgt)
    logger.error(msg)
    logger.error('Extra expected columns: {}, Extra received columns: {}'.format(expected_cols-data_has_cols,data_has_cols-expected_cols))
    raise IndexError(msg)

  return df,groups

def main():
  default_wb=config['workbook']
  parser = argparse.ArgumentParser(description ='Copies data from input file into tab "iande" or "current".')
  parser.add_argument('-s','--sheet',choices=['iande','current'],default='iande',help='which sheet - iande or current')
  parser.add_argument('-p','--path',default= None,help='The path and name of the input file. If not given will use "data/iande.tsv" or "data/iande_ytd.tsv" depending on sheet')
  parser.add_argument('-w','--workbook',default=default_wb,help=f'Target workbook. Default: {default_wb}')
  parser.add_argument('-f','--force', action='store_true', default=False, help='Use -f to ignore warning')
  
  args=parser.parse_args()
  path=args.path
  IANDE,CURRENT='iande.tsv','iande_ytd.tsv'
  if path is None:
    fn={'iande':IANDE,'current':CURRENT}[args.sheet]
    path=os.path.join('data',fn)
  ffy=config['first_forecast_year']
  table_info=config['sheets'][args.sheet]['tables'][0]
  data=read_iande_actl(data_info={'path':path})
  table='tbl_'+args.sheet
  data,fold_groups=prepare_iande_actl(workbook=args.workbook,target_sheet=args.sheet,df=data,force=args.force,f_fcast='Y%04d'%ffy)
  wkb = load_workbook(filename = args.workbook, read_only=False)
  data=forecast_formulas(table_info,data,ffy) # insert forecast formulas per config
  data=actual_formulas(table_info,data,ffy) # insert actual formulas per config
  wkb=fresh_sheet(wkb,args.sheet)
  wkb= write_table(wkb,target_sheet=args.sheet,df=data,table_name=table,groups=fold_groups)
  attrs=col_attrs_for_sheet(wkb,args.sheet,read_config())
  wkb=set_col_attrs(wkb,args.sheet,attrs)
  wkb.save(args.workbook)

if __name__ == '__main__':
  main()
