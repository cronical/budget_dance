#! /usr/bin/env python
"""Load data from 'data/iande.tsv' into tab 'iande_actl'

Handles category nesting as groups with subtotals

"""
import argparse
import pandas as pd
from utility import tsv_to_df,df_for_range,ws_for_table_name,df_for_range, get_val
from integrity import required_lines
from util.logs import get_logger

logger=get_logger(__file__)

parser = argparse.ArgumentParser(description ="Copies data from 'data/iande.tsv' into tab 'iande_actl' after doing some checks. ")
parser.add_argument("--force", "-f",action="store_true", help="Use -f to ignore warning")
args=parser.parse_args()

source='data/fcast.xlsm'
target = source
input_file="data/iande.tsv"
initialize_iande=False # don't initialize iande (only import into iande_actl)

# get the output of the Moneydance report
df=tsv_to_df (input_file,skiprows=3)
# some data clean ups
df.fillna(0, inplace=True) # replace the NaN's with zeros
df.query('Account!=0',inplace=True) # remove blank rows

#indent -other rows
def indent_other(str):
  if -1==str.find('-Other'):
    return str
  else:
    return '   '+str

# fix an anomoly in the md report
df.Account=df.Account.apply(indent_other)

# determine level of each row in the category hierarch
# hold the flat view in the key column for now
df.insert(loc=0,column='key',value=df.Account.str.lstrip()) # a temporary value - used to define level
#create a level indicator
# re-use the totals colum
df.rename(columns={'Total':'level'},inplace=True)
df['level']=((df.Account.str.len() - df.key.str.len()))/3
level=[int(x) for x in df.level.tolist()]
del df['level']

# create the real key field
def ravel(pathparts):
  path=''
  for pp in pathparts:
    if len(path)>0: path=path +':'
    path=path+pp
  return path

keys=[]
keyparts=df.Account.tolist()
rows=list(range(0,len(level))) # rows and columns are origin 0, excel uses origin 1 plus it contains the headings so for rows its off by 2
last_level=-1
for rw in rows:
  lev=level[rw]
  if 0==lev:pathparts=[]
  else:
    if lev > last_level: pathparts.append(keyparts[rw-1].strip())
    if lev < last_level: 
      pathparts=pathparts[:-1]
  a=pathparts.copy()
  a.append(keyparts[rw].strip())
  keys.append(ravel(a))
  last_level=lev

df['key']=keys

# change the year columns to Y+year format
for cn in df.columns.values.tolist():
  try:
    n=int(cn)
    nn = 'Y{}'.format(n)
    df.rename(columns={cn:nn},inplace=True)
  except ValueError:
    n=0

from openpyxl import load_workbook
wb = load_workbook(filename = source, read_only=False, keep_vba=True)
logger.info('loaded workbook from {}'.format(source))

# get the first forecast year from the general state table
ws=wb['utility']
ref=ws.tables['tbl_table_map'].ref
table_map = df_for_range(worksheet=ws,range_ref=ref)
table_name='tbl_gen_state'
ws_name =ws_for_table_name(table_map=table_map, table_name=table_name)
ws=wb[ws_name]
table=df_for_range(worksheet=ws,range_ref=ws.tables[table_name].ref)
f_fcst= get_val(table=table,line_key='first_forecast',col_name='Value')
print ("First forecast year is: %s"%f_fcst)

# make sure nothing of the forecast gets lost due to changed line.
req_lines, all_lines =required_lines(f_fcst)
if not req_lines.issubset(set(keys)):
  logger.warning("Existing forecast lines are not all present")
  logger.warning("The following line(s) exist in {} but do not occur in the file to import ({})"\
    .format(source,input_file))
  missing=list(req_lines.difference(set(keys)))
  for ms in missing: logger.info("   {}".format(ms))
  logger.info("Intialize iande flag is set to %r"%initialize_iande)
  if initialize_iande:
    print("That will remove those forecast lines from iande")
    if not args.force:
      logger.error("Exiting {}. Re-run with -f to continue".format(__file__))
      quit()
  else:
    logger.info("But we are only updating iande_actl, so we don't care.")
else:  
  logger.info("All forecast items are included in input file.")

# now report any new lines
new=list(set(keys).difference(all_lines))
new.sort()
if 0<len(new): 
  logger.info('Note: new lines (will be in iande_actl but not in iande, but you can manually add them there')
  logger.info('Note: lines with the suffix "-Other" occur when a transaction exists at a non-leaf in the category tree')
  logger.info('      If not desired, remove transaction and run again.  ')
for nw in new: logger.info("   {}".format(nw))




'''utility to aid with setting up groups'''
def getlev (e):
  return e[0]


from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


#set up to also seed the iande tab on the 2nd pass
sheets= ['iande_actl','iande']
tab_tgts=["tbl_iande_actl", 'tbl_iande']
enabled=[True,initialize_iande] # normally set to [True, False], but both True to initialize iande
tab_styles=['TableStyleMedium7','TableStyleMedium5']
# copy the data in to enabled tabs in file and into table, set up subtotals and groups

for ctl in zip(sheets,tab_tgts,enabled,tab_styles):
  if ctl[2]: #only if enabled
    sheet=ctl[0]
    tab_tgt=ctl[1]
    tab_style=ctl[3]

    from utility import fresh_sheet
    wb=fresh_sheet(wb,sheet)
    ws=wb[sheet]
    cols=list(range(0,df.shape[1]))

    seeding = tab_tgt=='tbl_iande' # seeding flags if this is for the iande tab
    seeding_col_names = []
    if seeding:
      end_year=2054
      last_imported=int(df.columns.tolist()[-1][1:])
      yr = last_imported+1
      while yr<=end_year:
        seeding_col_names.append('Y{}'.format(yr))
        yr=yr+1

    #transfer the dataframe into sheet

    # for set heading values for columns include any seeding extension
    all_col_names=df.columns.tolist() + seeding_col_names
    for cl in list(range(0,len(all_col_names))):
      ws.cell(1,cl+1,value=all_col_names[cl])
    
    # the data items - only for the imported data
    # for seeding of iande virtualize the actuals (all columns).
    # that is, refer to the values in the iande_actl tab
    for rw in rows:
      for cl in cols:
        if seeding and (cl > 1):
          val='=get_val($A{},"tbl_iande_actl",{}$1)'.format(rw+2,get_column_letter(cl+1))
        else:
          val=df.loc[df.index[rw]][cl]
        ws.cell(rw+2,cl+1,value=val)

    # set the label column width
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35

    # put in subtotal formulas for the numeric columns
    # format the numerics for the secion headings as blanks
    # the subtotals are aligned with the groups, so do those too
    cols= list(range(2,len(all_col_names)))

    #whenever the level goes down from the previous level its a total
    groups=[] # level, start, end
    last_level=-1
    for rw in rows:
      #for all cells apply formats
      for cl in cols:
        ws.cell(column=cl+1,row=rw+2).number_format='#,###,##0;-#,###,##0;"-"'
      level_change= level[rw]-last_level
      k=ws.cell(row=rw+2,column=1).value # get the key value from the file
      if level_change <0: #this should be a total line
        n=k.find(' - TOTAL') # see if it has the total label
        assert -1 != n # if not we are in trouble
        bare = k[:n]# # remove the total label
        try:
          x=keys.index(bare) # look for this in the keys - should be there
          #since we have the row for the section start - we'll blank out those zeros
          for cl in cols:
            ws.cell(column=cl+1,row=x+2).number_format='###'
          # prepare the grouping specs
          groups.append([level[rw]+1,x+2,rw+1])  
        except ValueError:
          assert False, '{} not found in keys'.format(k)
        
        # now replace the hard values with formulas
        # if its not a total just let the value stay there
        for cl in cols:
          let=get_column_letter(cl+1)
          formula='=subtotal(9,{}{}:{}{})'.format(let,x+2,let,rw+1)
          ws.cell(column=cl+1,row=rw+2, value=formula)
      
      if (-1!=k.find('TOTAL INCOME - EXPENSES')): # we are on the last row
        x=keys.index("Income - TOTAL")
        for cl in cols:
          let=get_column_letter(cl+1)
          formula='={}{}-{}{}'.format(let,x+2,let,rw+1)
          ws.cell(column=cl+1,row=rw+2, value=formula)  
      #for all cells apply formats
      last_level=level[rw]

    # set up the row groups highest level to lowest level
    # outline levels are +1 to our levels here
    # 0 here = 1 in Excel - such as Income, Expense
    # 1 here = 2 in Excel - such as I, T and X

    groups.sort(key=getlev)
    for grp in groups:
      ws.row_dimensions.group(grp[1],grp[2],outline_level=grp[0], hidden=grp[0]>2)

    rng=ws.dimensions
    tab = Table(displayName=tab_tgt, ref=rng)
    # Add a builtin style with striped rows and banded columns
    # The styles are seen on the table tab in excel broken into Light, Medium and Dark.
    # The number seems to be the index in that list (top to bottom, left to right, origin 1)
    style = TableStyleInfo(name=tab_style,  showRowStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    logger.info('table {} added'.format(tab_tgt))
    
    #open pyxl shape_writer getting a python future warning
    import warnings
    warnings.filterwarnings("ignore",category=FutureWarning)

    wb.save(target)
    logger.info('saved to {}'.format(target))

    # ws.column_dimensions.group('A','A', hidden=True)

