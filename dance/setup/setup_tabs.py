'''Utility to add tabs to the workbook and sets the tab colors

May be modified to add new tabs and re-run.  Unless overwrite is selected, does not replace or delete any tabs, only adds them
'''
import argparse
from json import JSONDecodeError 
from requests.exceptions import ConnectTimeout, ReadTimeout
from os.path import exists
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import Color

import pandas as pd
from numpy import nan
import yaml

from dance.util.books import col_attrs_for_sheet,set_col_attrs,freeze_panes,tab_color
from dance.util.logs import get_logger
from dance.setup.local_data import read_data, read_gen_state
from dance.util.files import read_config
from dance.util.row_tree import hier_insert,folding_groups,is_leaf,nest_by_cat,subtotal_formulas
from dance.util.tables import first_not_hidden,write_table,columns_for_table,conform_table
from dance.util.xl_formulas import actual_formulas,forecast_formulas, dyno_fields
from dance.setup.remote_data import request

def include_year(table_info,first_forecast_year,proposed_year):
  '''return True if proposed year should be displayed'''
  ao=False
  if 'actual_only' in table_info:
    ao=table_info['actual_only']
  if not ao:
    return True
  return first_forecast_year > proposed_year

def refresh_sheets(target_file,overwrite=False):
  '''create or refresh tabs'''
  logger=get_logger(__file__)
  config=read_config()
  years=range(config['start_year'],1+config['end_year'])
  wb=load_workbook(filename = target_file)
  
  sheets=wb.sheetnames
  logger.debug('{} existing sheets in {}'.format(len(sheets),target_file))

  for default_name in 'Sheet','Sheet1','Sheet2','Sheet3':
    if default_name in sheets:
      del wb[default_name]
      logger.debug('Removed "{}"'.format(default_name))

  table_map={} # build list of tables and which sheet they are on
  ffy=config['first_forecast_year']
  general_state=read_gen_state(config)
  gs_ffy=general_state['first_forecast']
  gs_ffy['Value']=f'Y{ffy}' # prefer the value from the config
  _=general_state # it is used by referencing locals(), but pylance can't figure it out

  for sheet_group,sheet_group_info in config['sheet_groups'].items():
    for sheet_name,sheet_info in config['sheets'].items():
      if sheet_info['sheet_group'] != sheet_group:
        continue
      color=sheet_group_info['color']
      existing=False
      try:
        _=sheets.index(sheet_name) # is it already there?
        ws = wb[sheet_name]
        existing=True
      except ValueError:
        ws=wb.create_sheet(sheet_name)
        logger.debug('sheet {} added'.format(sheet_name))
      if existing:
        if not overwrite:
          logger.info('Sheet {} already exists, so not initializing'.format(sheet_name))
          continue
        else:
          del wb[sheet_name]
          ws=wb.create_sheet(sheet_name)
          logger.debug('sheet {} deleted and recreated'.format(sheet_name))

      ws.sheet_properties.tabColor= tab_color(color)

      table_location=1 # add to this for each subsequent table
      for table_info in sheet_info['tables']:

        #take specified or default items
        key_values={'title_row':table_location,'start_col':1,'include_years':False}
        for k in key_values:
          if k in table_info:
            key_values[k]=table_info[k]

        title_cell=f'{get_column_letter((key_values["start_col"]-1)+first_not_hidden(table_info))}{key_values["title_row"]}'
        ws[title_cell].value=table_info['title']
        ws[title_cell].font=Font(name='Calibri',size=16,color='4472C4')

        col_def=columns_for_table(wb,sheet_name,table_info['name'],config)
        col_count=col_def.shape[0]
        groups=None
        # Excel seems to need at least one data row, so a row of blanks if no data is provided
        if 'data' not in table_info:
          data=pd.DataFrame([[None]*col_count],columns=col_def.name)  
        else: # if table has a data source add those rows
          assert 'source' in table_info['data']
          data_info=table_info['data']
          valid_sources=['internal','remote','local','none']
          source=data_info['source']
          if source not in valid_sources:
            logger.error('Only the following are valid data sources {}'.format(valid_sources))
            quit()
          if source == 'none':
            data=pd.DataFrame(columns=col_def.name)                  
            if 'hier_separator' in data_info: # this will want groups, folding, subtotals
              # for now only for sourec = none. 
              # same logic exists in iande_actl_load, taxes_load which may be able to move

              data=nest_by_cat(data,cat_field='Line') # creates key, leaf and level fields
              data=hier_insert(data,table_info,sep=data_info['hier_separator']) # insert any specified lines into hierarchy
              data=is_leaf(data)# mark rows that are leaves of the category tree
              data,groups=folding_groups(data)
              del data['is_leaf'] # clear out temp field
              title_row=1
              if 'title_row' in table_info:
                title_row=table_info['title_row']
              data=subtotal_formulas(data,groups)
          if source == 'internal':
            try:
              var_name=data_info['name']
              data=locals()[var_name]
            except KeyError:
              logger.error('configured internal data source {} does not exist.'.format(var_name))
              quit()
            if not isinstance(data,dict):
              logger.error('configured internal data source {} is not a dict.'.format(var_name))
              quit()
          if source == 'remote':
            if 'api_key' in data_info:
              fn='./private/api_keys.yaml'
              if not exists(fn):
                logger.error('call for remote data but file {} does not exist'.format(fn))
                quit()
              with open(fn) as y:
                api_keys=yaml.load(y,yaml.Loader)
              key_name=data_info['api_key']
              if key_name not in api_keys:
                logger.error('call for remote data but api key name {} does not exist in {}'.format(key_name,fn))
                quit()
              data_info['api_key']=api_keys[key_name]
              logger.debug('API key retrieved from private data')
            try:
              data=request(table_info)
              logger.debug('pulled data from remote')
            except (JSONDecodeError, ValueError, ConnectTimeout, ReadTimeout) as e:
              logger.error('*** Remote data problem, possible maintenance window')
              logger.error(str(e))
              logger.error('*** Skipping table: %s'%table_info['name'])
              continue
          if source=='local': 
            data,groups=read_data(data_info,years,ffy,target_file=target_file,table_map=table_map,
            title_row=table_info.get('title_row',1))
          if isinstance(data,dict):
            data=pd.DataFrame.from_dict(data,orient='index')
            data=data.reset_index()
            data.columns=list(col_def.name)
          if isinstance(data,list):
            data=pd.DataFrame(list,columns=col_def.name)
          if isinstance(data,pd.DataFrame):
            # special case for YTD which gets as_of_date as new column heading 
            if 'YTD' in col_def.name.tolist():
              col_def.loc[col_def.name=='YTD',['name']]=data.columns[3] #TODO  remove hardcoded 3
            # it may be that the 1st column is in the index,
            # and/or there may be a new year 
            # so fix that
            mismatch=list(set(col_def.name)-set(data.columns))
            data.reset_index(inplace=True)
            for missing in mismatch:
              col_no=list(col_def.name).index(missing)
              if col_no==0:
                data=data.rename(columns={'index':missing})
              else:
                data[missing]=nan
        

        
        data=dyno_fields(table_info,data)# any dynamic field values
        data=conform_table(data,col_def['name'])
        data=forecast_formulas(table_info,data,ffy,wb=wb,table_map=table_map) # insert forecast formulas per config
        data=actual_formulas(table_info,data,ffy,wb=wb,table_map=table_map) # insert actual formulas per config
        edit_checks=None
        if 'edit_checks' in table_info:
          edit_checks=table_info['edit_checks']
        wb=write_table(wb=wb,target_sheet=sheet_name,table_name=table_info['name'],df=data,groups=groups,title_row=key_values['title_row'],edit_checks=edit_checks,table_map=table_map)
        attrs=col_attrs_for_sheet(wb,sheet_name,config)
        wb=set_col_attrs(wb,sheet_name,attrs)
        wb=freeze_panes(wb,sheet_name,config)
        # save after each sheet to allow successive sheets to locate earlier sheet data
        wb.save(target_file)
        logger.debug('workbook {} saved'.format(target_file))
        table_map[table_info['name']]=sheet_name
        table_location += 3+data.shape[0]
      ws.sheet_view.zoomScale=config['zoom_scale']

  wb.save(filename=target_file)
  logger.info('workbook {} saved'.format(target_file))

if __name__=='__main__':
  # execute only if run as a script
  parser = argparse.ArgumentParser(description ='set tabs the given workbook')
  parser.add_argument('target_file', help='provide the name of the output file')
  args=parser.parse_args()
  refresh_sheets(args.target_file)
