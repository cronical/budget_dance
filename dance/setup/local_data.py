'''get annual various local data'''
import pandas as pd
from openpyxl.utils import get_column_letter
from dance.iande_actl_load import read_iande_actl, prepare_iande_actl
from dance.transfers_actl_load import read_transfers_actl, prepare_transfers_actl
from dance.util.files import tsv_to_df
from dance.util.tables import this_row
from dance.util.logs import get_logger
logger=get_logger(__file__)

def read_data(data_info,years=None,ffy=None,target_file=None):
  '''Read data and prepare datafor various tabs and prepare it for insertion into the workbook.
  Uses the type value in data_info to determine what to do.

  args:
    data_info: a dictionary that contains data to control the reading and preparation
    years: Some types require additional info such as the years for the column.
    ffy: first forecast year as integer
    target_file: supports the case when there is a need to look at previously written worksheets

  returns: dataframe with columns specific to the type and groups (which may be None)

  raises: ValueError when things don't make sense.
    Such as in the case of balances if there is an account not found in the accounts worksheet
    '''
  groups=None
  if data_info['type']== 'md_acct':
    df= read_accounts(data_info)
    df= prepare_account_tab(data_info,df)
  if data_info['type']=='md_bal':
    df=read_balances(data_info,target_file)
    df=prepare_balance_tab(years,ffy,df)
  if data_info['type']=='md_iande_actl':
    df=read_iande_actl(data_info=data_info)
    df,groups=prepare_iande_actl(workbook=target_file,target_sheet=data_info['sheet'],df=df)
  if data_info['type']=='md_transfers_actl':
    df=read_transfers_actl(data_info=data_info)
    df,groups=prepare_transfers_actl(workbook=target_file,df=df,f_fcast=ffy)
  return df,groups

def filter_nz(df,include_zeros):
  '''get a filter for the the rows in a dataframe that have a balance or are allowed by config.

  args:
    df: a dataframe with fields, Current Balance and Account
    include_zeros: a list of account names that should be kept even if zero

  returns: boolean series, where True indicates item should be kept.
  '''
  nz=df['Current Balance']!=0
  za=df['Account'].apply(lambda x: x in include_zeros)
  nzza=pd.concat([nz, za],axis=1)
  return nzza.any(axis=1)

def no_suffix(acct_total_key):
  '''Drop the total sufix
  args:
    acct_total_key: string including " - Total" as a suffix
  returns: same without the suffix
  '''
  return ' - '.join(acct_total_key.split(' - ')[:-1])

def read_accounts(data_info):
  '''Get the account data as specified in the data_info.

  Depending on the data_info may take items at a total level or leaf level.
  All investment accounts are processed at the total level for each account.  Securities are dropped.

  args:
    data_info: a dictionary with the path of the source file, and three lists:
    1. group - identifies categories or total levels (with subaccounts)
    2. no_details_for - identifies categories that are covered by total levels
    3. include_zeros - indicates accounts that are wanted even though they are zero balance (otherwise ignored)

  returns: dataframe including 'Account', 'Type', 'Current Balance', is_total, and 'level'
  '''
  # doing the same with data frame
  types={'Assets':'A','Bank Accounts':'B','Credit Cards':'C','Investment Accounts':'I','Liabilities':'L','Loans':'N'}
  groups=data_info['group']
  no_details_for=data_info['no_details_for']
  include_zeros=set(data_info['include_zeros'])
  include_zeros.update(groups)# add groups to the include zeros list
  g_totals=[g+' - Total' for g in groups]

  df=tsv_to_df(data_info['path'],skiprows=3)
  logger.info('{} rows read'.format(len(df)))
  df=df.dropna(how='any',subset='Account') # remove blank rows
  df.reset_index(drop=True,inplace=True)
  logger.info('{} non-blank rows'.format(len(df)))
  df['Account']=df['Account'].str.strip()

  # establish category for each row
  df['category']=''
  cat='Bank Accounts'# should be in the first row, but this keeps it clean
  for ix,row in df.iterrows():
    if pd.isnull(row['Notes']):
      cat=row['Account']
      df.loc[ix,'Notes']='' # to allow selecting securities later
    df.loc[ix,'category']=cat

  # build a boolean grid to track where each account is handled
  attr=['is_total','keep']
  dispo=['is_heading','rollup','zero_rule','by_total','no_detail','extra_totals']
  for col in attr+dispo:
    # is_total - keeps track of totals
    # keep - indicates the ones to keep
    # The rest are dispositions
    # is_heading - well, its a heading
    # rollup - is when a total is selected
    # non zero rule - is True the zero rule was applied
    # by_total - is True if its handled by a total
    # no_detail - is True if its excluded by config
    # extra_totals - grand total, total of all investments and total of category in name conflict
    df[col]=False

  # set the extra totals flag
  xt=[x+ ' - Total'for x in no_details_for]+['Total']
  df.loc[df['Account'].isin(xt),'extra_totals']=True
  logger.info('{} rows marked as extraneous totals'.format(len(xt)))

  tot_rows=df['Account'].str.contains(' - Total') # all the total rows on the report
  logger.info('{} totals identified'.format(tot_rows.sum()))

  # flag the investment total rows except total for all investments.
  i_tots=tot_rows & (df['category']=='Investment Accounts') & (df['Account'] !='Investment Accounts - Total')
  i_totals=df.loc[i_tots,'Account'].tolist()
  logger.info('{} investment accounts identified'.format(len(i_totals)))

  # establish the account level by use of the headings and totals
  df['level_change']=0
  for ix,row in df.loc[tot_rows,['Account']].iterrows():
    acct_total_key=row['Account']
    acct_key=' - '.join(acct_total_key.split(' - ')[:-1])
    # mark the heading and the total lines
    sel=df['Account'].apply(lambda x: x in [acct_key , acct_total_key])
    ixs= df.loc[sel].index.tolist()
    for i,ix in enumerate(ixs):
      delta=[1,-1][i>=len(ixs)/2]
      df.loc[ix,'level_change']=delta
      if delta==-1:
        df.loc[ix,'is_total']=True
  df['level']=df['level_change'].cumsum(axis=0)
  del df['level_change']
  logger.info('assigned account hierarchy levels. Max level is {}'.format(df.level.max(axis=0)))

  #remove the suffix from totals
  df.loc[tot_rows,'Account']=df.Account.apply(no_suffix)
  df.loc[tot_rows,'is_total']=True

  # work with the ranges between the headings and the totals
  dups=[]
  for ix,row in df.loc[tot_rows,['Account']].iterrows():
    acct_key=row['Account']
    if acct_key in dups:
      continue
    acct_total_key=acct_key + ' - Total'
    sel=df['Account']== acct_key     # mark the heading and the total lines

    if sum(sel)==4:# first resolve potential name conflict between category and sub-account total
      # by removing the category instance of the name
      logger.info('naming conflict identified on name: {}'.format(acct_key))
      sa=(df.loc[sel,'level']==0) # flag the category total line
      sal=list(sa.values)
      df.loc[sa.index[0],'is_heading']=True
      df.loc[sa.index[-1],'extra_totals']=True # mark disposition of the one we will ignore
      logger.info('ignoring the one at level {}'.format(df.loc[sa.index[-1],'level']))
      sal=sal[len(sa)//2:len(sa)]# the 2nd half of the list.
      sal=list(reversed(sal))+sal # reflect across the midpoint so as to also mark the header
      sel.loc[sa[sal].index]=False # turn them off
      dups+=[acct_key]
    assert sum(sel)==2, 'too many name conflicts on {}'.format(acct_key)

    # now that there is no name conflict mark dispostion of the rows
    start,end=tuple(df.loc[sel,[]].index)
    df.loc[[start],'is_heading']=True # no further use for headings, just mark it for quality check
    ixs=range(start+1,end) # all the items included in this total
    if acct_total_key in g_totals+i_totals: # process all investment accounts and the configured groups
      df.loc[ixs,'by_total']=True
      df.loc[ixs,'keep']=False # leaf nodes covered by a total
      nzr=filter_nz(df.loc[[end]],include_zeros) # keep this total per the non-zero rules
      df.loc[[end],'zero_rule']=True
      df.loc[[end],'rollup']=nzr
      df.loc[[end],'keep']=nzr
    else: # items not investments or configured groups
      if acct_key in no_details_for: #specifically excluded by config
        this_lev=df.loc[ixs,'level']==df.loc[start,'level'] # only go one level, deeper needs its own rule
        not_tot=~df.loc[ixs,'is_total']
        mark=not_tot & this_lev
        df.loc[[x for x in ixs if mark.loc[x]],'no_detail']=True
      else:
        nzr=filter_nz(df.loc[ixs],include_zeros)
        df.loc[ixs,'keep']=nzr# the leaf nodes we will keep unless we have a total at a higher level, in which case, they will later be overwritten
        df.loc[ixs,'zero_rule']=True
        df.loc[end,'is_total']=True
        nzr=filter_nz(df.loc[[end]],include_zeros) # keep this total per the non-zero rules
        df.loc[[end],'zero_rule']=True
        df.loc[[end],'rollup']=nzr
        df.loc[[end],'keep']=nzr

  logger.info('Rows counts for each type of disposition:')
  for col in dispo:
    logger.info('{}: {}'.format(col,df[col].sum(axis=0)))

  flag_tot=df[['is_heading','rollup','zero_rule','by_total','no_detail','extra_totals']].sum(axis=1)
  if flag_tot.min()==0:
    logger.warning('The following accounts were not disposed of properly')
    for _,row in df.loc[flag_tot==0].iterrows():
      logger.warning(row['Account'])

  df=df.loc[df.keep]
  df['Type']=[types[x] for x in df.category.tolist()]
  logger.info('Returning {} rows'.format(len(df)))
  return df[['Account','Type','Current Balance','is_total' ,'level']]

def read_balances(data_info,target_file):
  bal_df=read_accounts(data_info)
  acct_df=pd.read_excel(target_file,sheet_name='accounts',skiprows=1)
  if len(bal_df)!=len(bal_df.merge(acct_df,on='Account')): # make sure all the balances are in the account list
    raise ValueError('Balance account(s) are not all in the Accounts table')
  logger.info('All balance accounts are in the accounts table.')
  return bal_df

def prepare_account_tab(data_info, in_df):
  '''Add in the added fields for the account worksheet
  args:
    data_info: a dictionary describing the data (from the config.yaml file)
    in_df: a dataframe with columns: Account, Type, Current Balance, is_total and level

  returns: a dataframe for the accounts tab
  '''
  df=in_df
  tax_free_keys=[]
  if 'tax_free_keys'in data_info:
    tax_free_keys=data_info['tax_free_keys']
  account_names=df.Account.tolist()
  df['Income Txbl']= [ int(not any(y in x for y in tax_free_keys)) for x in account_names]
  df['Active']=(df[['Current Balance']] != 0).astype(int) # default zero accounts to inactive
  df['Rlz share']=0
  df['Unrlz share']=1
  acct_ref='= '+this_row('Account')
  source_formulas=[ acct_ref,acct_ref + ' & " - TOTAL"' ]  #formula for (sub-accounts),  (leaves, and categories)
  flag=(df.level>0) & (df.Type!='I') & (df.is_total)
  df['Actl_source']=[source_formulas[ x] for x in flag]
  source_tabs=['tbl_transfers_actl','tbl_invest_actl']# actuals are sourced from this, (depends on account type)
  df['Actl_source_tab']= [source_tabs[x=='I']for x in df.Type.tolist()]
  df['Fcst_source']='zero' # a default, but should be changed by user
  del df['is_total']
  del df['level']
  return df

def prepare_balance_tab(years,first_forecast,in_df):
  '''Add in the added fields for the balance worksheet
  args:
    years: iterable with the numeric years to be appended to the columns
    first_forecast: integer year that is the 1st to use the forecast formulas
    in_df: a dataframe with at least columns: Account, and Current Balance

  returns: a dataframe for the balance tab
  '''

  acct_ref=this_row('AcctName')
  key_ref=this_row('Key')
  repeated_formulas={
    'Key':'=@CONCATENATE( {},{})'.format(this_row('ValType'), acct_ref),
    'Type':'=@get_val({},"tbl_accounts",D$2)'.format(acct_ref),
    'Income Txbl': '=@get_val( {},"tbl_accounts",E$2)'.format(acct_ref),
    'Active': '=@get_val( {},"tbl_accounts",F$2)'.format(acct_ref)
  }
  lead_cols=['Key','ValType','AcctName','Type','Income Txbl','Active']

  # the actual and the forecast formulas
  # braces are filled in before inserting into the worksheet
  actl_formulas={
    'Rate':'=simple_return( %s,{}$2)' % (acct_ref),
    'Start Bal':'=get_val("End Bal" &  %s,"tbl_balances",{}$2)' % (acct_ref),
    'Add/Wdraw':'=-add_wdraw( %s,{}$2)' % (acct_ref),
    'Rlz Int/Gn':'=@gain( %s,{}$2,TRUE)' % (acct_ref),
    'Unrlz Gn/Ls':'=@gain( %s,{}$2,FALSE)' % (acct_ref),
    'End Bal': '=@endbal( %s,{}$2)'  % (acct_ref) }
  fcst_formulas={
    'Rate':'=rolling_avg("tbl_balances", %s,{}$2,3)' % key_ref,
    'Start Bal':'=get_val("End Bal" &  %s,"tbl_balances",{}$2)' % (acct_ref),
    'Add/Wdraw':'=-add_wdraw( %s,{}$2)' % (acct_ref),
    'Rlz Int/Gn':'=@gain( %s,{}$2,TRUE)' % (acct_ref),
    'Unrlz Gn/Ls':'=@gain( %s,{}$2,FALSE)' % (acct_ref),
    'End Bal': '=@endbal( %s,{}$2)' % (acct_ref)  }

  val_types=actl_formulas.keys()
  r_count=len(val_types)
  df=pd.DataFrame([])
  # prepare the data for each account
  for _,row in in_df.iterrows():
    a_df=pd.DataFrame([])
    for c in lead_cols:
      if c in repeated_formulas:
        a_df[c]=[repeated_formulas[c]]*r_count
      else:
        if c== 'ValType':
          a_df[c]=val_types
        else:
          a_df[c]=[row['Account']]*len(val_types)
    for c in years:
      formulas=[]
      for vt in val_types:
        raw_formula=[actl_formulas,fcst_formulas][c >= first_forecast][vt]
        yx=years.index(c)
        this_year_col=get_column_letter(1+len(lead_cols)+yx)
        if vt == 'Start Bal': # only type to use prior year col
          if yx >0:
            prior_year_col=get_column_letter(len(lead_cols)+yx)
            formula=raw_formula.format(prior_year_col)
          else: # initial year is the balance from the input
            formula=float(in_df.loc[in_df.Account == row['Account'],'Current Balance'])
        else:
          formula=raw_formula.format(this_year_col)
        formulas+=[formula]
      a_df['Y{}'.format(c)]=formulas
    df=pd.concat([df,a_df],axis=0)
  df.reset_index(inplace=True,drop='True')
  return df
