'''experiment'''
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl import load_workbook
fn='data/test_wb.xlsx'
wb=load_workbook(fn)
ws=wb['taxes']
dxf= DifferentialStyle(font=Font(bold=True),fill=PatternFill(start_color='EE1111',end_color='EE1111'))
rule=Rule(type='expression',dxf=dxf,stopIfTrue=False)
rule.formula=['$A3="Adjusted Gross"']
ws.conditional_formatting.add('A3:U102',rule)
cflist=ws.conditional_formatting
wb.save(fn)
