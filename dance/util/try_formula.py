'''Discover logic to set array formula'''

from openpyxl.worksheet.formula import ArrayFormula
from dance.util.tables import load_workbook, df_for_table_name
from dance.util.xl_formulas import filter_parser,eval_criteria

source='data/test_wb.xlsm'
table_name='tbl_transfers_plan'
ws_name='transfers_plan'

ref_table='tbl_accounts'
ref_ws='accounts'
formula= '=FILTER(tbl_accounts[Account],(tbl_accounts[Active])*(tbl_accounts[No Distr Plan]))'

ref_df=df_for_table_name(ref_table,workbook=source,data_only=True)
table,field,criteria=filter_parser(formula)
sel=eval_criteria(criteria,ref_df)
ref_df.index.name=field
ref_df=ref_df.loc[sel].reset_index()
base=3
col='F'

# turns out you don't need to populate the array
#wb = load_workbook(filename = source, read_only=False, keep_vba=True,data_only=True)
#ws=wb[ws_name]
#for ix,row in ref_df.iterrows():
#    ws[col+str(base+ix)]= row[field]
#wb.save(source)
#print('saved values')

wb = load_workbook(filename = source, read_only=False, keep_vba=True)
ws=wb[ws_name]

# needs the prefixes - otherwise excel complains and strips out formula
formula=formula.replace('FILTER','_xlfn._xlws.FILTER')

# if the size of ref is less than the size of the result of filter, it gets truncated.
# if its is greater then its filled with #N/A
# this appears to work but it does wrap the visible formula with curly braces
ref='%s%d:%s%d'%(col,base,col,base+ref_df.shape[0]-1)
ws[col+str(base)]=ArrayFormula(ref,formula)

# if dynamic array formula is manually entered, we get cm which is an index into metadata, 
# which also has to exist and the schemas need to allow it:
#       <c r="F3" s="13" t="str" cm="1">
#       <f t="array" ref="F3:F24">_xlfn._xlws.FILTER(tbl_accounts[Account],tbl_accounts[Active]=1)</f>
#       <v>Bank Accounts</v>
#     </c>

# created by the openpyxl formula_attributes resulting in curly braces
#       <c r="F3" s="13">
#        <f t="array" ref="F3:F24">_xlfn._xlws.FILTER(tbl_accounts[Account],tbl_accounts[Active]=1)</f>
#        <v></v>
#      </c>

wb.save(source)
print('saved formula')
