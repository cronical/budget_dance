'''Workbook oriented utilities'''
from openpyxl.utils.cell import get_column_letter

from dance.util.logs import get_logger
from dance.util.tables import columns_for_table
from dance.util.ui import tab_color
  
def fresh_sheet(wb,sheet_name,color='58BD2D'):
  '''Create or refresh a worksheet in a workbook.

  Removes the named sheet (if it exists) from the workbook and created a fresh one at the same location (or at the end)
  If not already existing will create the new sheet at the end. Copies tab config_color if the sheet already exists.

  args:
    wb: An openpyxl workbook
    sheet_name: The name of the worksheet
    color: Only used if the worksheet does not already exist. 
    Either index into theme pallette or and rgb color string.

  Returns: the workbook
  '''
  logger=get_logger(__file__)
  ix=len(wb.sheetnames)
  if sheet_name in wb.sheetnames:
    ix= wb.sheetnames.index(sheet_name)
    ws = wb[sheet_name]
    tc=ws.sheet_properties.tabColor
    wb.remove(ws)
    logger.info('deleted worksheet {}'.format(sheet_name))
  else:
    tc=tab_color(color)
  ws=wb.create_sheet(sheet_name,ix)
  logger.info('created worksheet {}'.format(sheet_name))
  ws.sheet_properties.tabColor=tc
  return wb

def col_attrs_for_sheet(wb,sheet,config):
  '''Get the column width and hidden attributes for a sheet based on the config.
  Handles 1 or more tables per worksheet.
  If more than one table uses a column, the first one wins.
  Used to project the columns onto the worksheet

  args:
    wb: the workbook
    sheet: the name of the worksheet
    config: the full config as a dict

  returns: dictionary where the key is the excel column number and the value is the width
  '''
  attrs={}
  sheet_info=config['sheets'][sheet]
  for table_info in sheet_info['tables']:
    df=columns_for_table(wb,sheet,table_info['name'],config)
    start_col=1
    if 'start_col' in table_info:
      start_col=table_info['start_col']
    for ix,rw in df.iterrows():
      cn=ix+start_col
      if cn not in attrs:
        attrs[cn]={'width':rw['width'],'hidden':rw['hidden']}
  return attrs

def set_col_attrs(wb,sheet,attrs):
  '''set column attributes such as width and hidden

  args:
    wb: the openpyxl workbook
    sheet: the worksheet name
    attrs: a dictionary with key = excel column number and value another dict with the attributes

  returns: the workbook
  '''
  ws=wb[sheet]
  for k,v in attrs.items():
    ws.column_dimensions[get_column_letter(k)].width = v['width']
    ws.column_dimensions[get_column_letter(k)].hidden = v['hidden']
  return wb

def freeze_panes(wb,sheet,config):
  '''Set a freeze pane point for sheet based on the config.
  Only operates on the first table of a sheet and only if include_years is True.

  args:
    wb: the workbook
    sheet: the name of the worksheet
    config: the full config as a dict

  returns: a possibly modified workbook
  '''
  table_info=config['sheets'][sheet]['tables'][0]
  if not table_info['include_years']:
    return wb
  tr=2
  if 'title_row' in table_info:
    tr=table_info['title_row']
  sc=1
  if 'start_col' in table_info:
    sc=table_info['start_col']
  x=len(table_info['columns'])
  let=get_column_letter(x+sc)
  address=let+'{}'.format(1+tr)
  wb[sheet].freeze_panes=address
  return wb
