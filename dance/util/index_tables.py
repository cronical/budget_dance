"""Utility to map all the tables in the workbook to their worksheets

Gathers all the tables names across all worksheets
and cross references them to their worksheets
Stores the results in the utility tab.
This is needed to allow the code in the worksheet to only know the table names.
"""

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
import pandas as pd
from util.logs import get_logger

def main():
  logger=get_logger(__file__)

  target='data/test_wb.xlsx'
  wb=load_workbook(filename = target)
  sheets=wb.worksheets
  data = []

  for ws in sheets:
    if ws.title != 'utility':
      tabs = ws.tables.keys()
      for t in tabs:
        data.append([t,ws.title])
    else:
      util_ws=ws

  df = pd.DataFrame(data, columns = ['Table', 'Worksheet'])

  #transfer the dataframe into sheet
  ws=util_ws
  tab_tgt='tbl_table_map'

  cols=list(range(0,df.shape[1]))
  rows=list(range(0,df.shape[0]))
  for cl in cols:
    ws.cell(1,cl+1,value=df.columns[cl])
  for rw in rows:
    for cl in cols:
      ws.cell(rw+2,cl+1,value=df.loc[df.index[rw]][cl])

  #remove any left from prior run
  for tab in ws.tables:
    if tab==tab_tgt:
      del ws.tables[tab_tgt]
      break

  rng='A1:{}{}'.format(get_column_letter(len(cols)),1+len(rows))
  tab = Table(displayName=tab_tgt, ref=rng)
  ws.add_table(tab)

  wb.save(target)
  logger.info(f'Information about {df.shape[0]} tables written to {target}')

if __name__=='__main__':
  main()
