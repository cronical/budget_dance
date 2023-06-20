'''Logic to set data validation'''

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from dance.util.tables import df_for_range,df_for_table_name,load_workbook,ws_for_table_name

source='data/test_wb.xlsm'
table_name='tbl_transfers_plan'
columns=['From_Account','To_Account']
wb = load_workbook(filename = source, read_only=False, keep_vba=True)
utility=wb['utility']
table_map = df_for_range(worksheet=utility,range_ref=utility.tables['tbl_table_map'].ref)
ws_name =ws_for_table_name(table_map=table_map, table_name=table_name)
ws=wb[ws_name]

# replace this reference with a filter
filter= '=SORT(FILTER(tbl_accounts[Account],(tbl_accounts[Active]=1)*(tbl_accounts[No Distr Plan]=1)))'
filter_loc='F3'


accounts=df_for_table_name(table_name='tbl_accounts',workbook=source)
choices=accounts.loc[(accounts['No Distr Plan']==1) & (accounts['Active']==1)].index.to_list()
choices=','.join(choices)
choices='"'+choices+'"'
choices="=accounts!$A$3:$A$32"


ref=ws.tables[table_name].ref
df=df_for_range(worksheet=ws,range_ref=ref).reset_index()
col_refs=[get_column_letter(1+list(df.columns).index(a)) for a in columns ]
row_refs=[coordinate_from_string(a)[1] for a in ref.split(':')]
row_refs[0]+=+1 # go to row after heading
ws.data_validations.dataValidation.clear() # remove prior items
dv = DataValidation(type="list", formula1=choices, allow_blank=True) #'"Bank Accounts,Credit Cards,Bat - Mobile - 1"'
for col in col_refs:
    cells='%s%d:%s%d'%(col,row_refs[0],col,row_refs[1])
    dv.add(cells)
ws.add_data_validation(dv)
wb.save(source)

pass