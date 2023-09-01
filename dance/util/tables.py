'''Utilities dealing with worksheet tables.'''
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting import Rule
from openpyxl.styles import Font, Alignment, PatternFill, Side,Border
from openpyxl.styles.numbers import BUILTIN_FORMATS,FORMAT_PERCENTAGE_00,FORMAT_NUMBER
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableStyleInfo

from dance.util.logs import get_logger
from dance.util.files import read_config
from dance.util.sheet import df_for_range
from dance.util.xl_eval import filter_parser,eval_criteria
from dance.util.xl_formulas import is_formula,prepare_formula

logger=get_logger(__file__)

def ws_for_table_name(table_map=None,table_name=None):
  '''Return name of worksheet holding table given a data frame holding the table'''

  return table_map.loc[table_name,'Worksheet']

def table_as_df(wb,table_name):
  '''converts an table_name into a dataframe with the 1st column as index, and the range it occupies
  args: wb = an open workbook in memory
        table name = an existing table

  requires the table map to exist in the workbook

  returns the dataframe, the worksheet name, the range that it occupies
  '''

  ws=wb['utility']
  ref=ws.tables['tbl_table_map'].ref
  table_map = df_for_range(worksheet=ws,range_ref=ref)
  ws_name =table_map.loc[table_name,'Worksheet']
  ws=wb[ws_name]
  ref=ws.tables[table_name].ref
  table=df_for_range(worksheet=ws,range_ref=ref)

  return table, ws_name,ref

def df_for_table_name(table_name=None, workbook='data/fcast.xlsm',data_only=False,table_map=None):
  '''Opens the file, and extracts a table as a Pandas dataframe

  args:
    table_name:
    workbook:
    data_only: True to get the data not the formula
    table_map: in case the utility tab is not yet available, provide a dict mapping tables to worksheets

  returns: a dataFrame with the first column as the dataframe index

  raises:
    ValueError: when the file does not exist or its structures are not right

  '''
  try:
    wb = load_workbook(filename = workbook, read_only=False, keep_vba=True, data_only=data_only)
    if table_map is None:
      ws=wb['utility']
      ref=ws.tables['tbl_table_map'].ref
      table_map = df_for_range(worksheet=ws,range_ref=ref)
    else: # convert working form to stored form
      table_map=pd.DataFrame([(k,v) for k,v in table_map.items()],columns=['Table','Worksheet'])
      table_map.set_index('Table',inplace=True)
    ws_name =ws_for_table_name(table_map=table_map, table_name=table_name)
    ws=wb[ws_name]
    table=df_for_range(worksheet=ws,range_ref=ws.tables[table_name].ref)
  except (FileNotFoundError,KeyError) as error:
    raise ValueError('workbook({}), worksheet({}) or internal structures not available'.format(workbook,table_name)) from error
  logger.debug('Read table {} from {}'.format(table_name,workbook))
  logger.debug('  {} rows and {} columns'.format(table.shape[0],table.shape[1]))
  return table

def df_for_table_name_for_update(table_name=None):
  '''Opens the file, and extracts a table as a dataFrame with the first column as the dataframe index
  returns dataframe, worksheet, range_ref and workbook
  since its going to be updated don't allow data only '''
  source='data/fcast.xlsm'
  wb = load_workbook(filename = source, read_only=False, keep_vba=True)
  logger.info('Loaded workbook from {}'.format(source))
  ws=wb['utility']
  ref=ws.tables['tbl_table_map'].ref
  table_map = df_for_range(worksheet=ws,range_ref=ref)
  ws_name =ws_for_table_name(table_map=table_map, table_name=table_name)
  ws=wb[ws_name]
  table=df_for_range(worksheet=ws,range_ref=ws.tables[table_name].ref)
  logger.info('{} rows and {} columns in sheet {}'.format(table.shape[0],table.shape[1],ws))
  return table, ws,ws.tables[table_name].ref,wb

def get_value(frame, line_key ,  col_name):
  '''get a single value from a dataframe'''
  return frame.loc[line_key,col_name]

def put_val(frame, line_key ,  col_name, value):
  '''Put a single value into a dataframe'''
  frame.loc[line_key,col_name]=value

def get_value_for_key(wb,key):
  '''Get a value from the general state table

  args:
    wb: the openpyxl workbook
    key: an item in the general (state) table

  returns: the value from the generate (state) table

  raises:
    ValueError: in case the structure in the file is not yet in place or the key is missing.
    Oddly ValueError works but KeyError does nothing.
  '''

  try:
    ws=wb['utility']
    ref=ws.tables['tbl_table_map'].ref
    table_map = df_for_range(worksheet=ws,range_ref=ref)
    table_name='tbl_gen_state'
    ws_name =ws_for_table_name(table_map=table_map, table_name=table_name)
    ws=wb[ws_name]
    table=df_for_range(worksheet=ws,range_ref=ws.tables[table_name].ref)
    value= get_value(table,line_key=key,col_name='Value')
    return value
  except KeyError as err:
    raise ValueError(str(err)) from None

def get_f_fcast_year(wb,config):
  '''get the first forecast year, using the value from the workbook if available, otherwise from the config

  args:
    wb: the workbook
    config: the config dict

  returns: integer year which is the first forecast
  '''
  ffy=None
  try:
    ffy=get_value_for_key(wb,'first_forecast')
    ffy=int(ffy[1:])
  except ValueError:
    ffy=config['first_forecast_year']
  return ffy

def first_not_hidden(table_info):
  '''determine the first column that is not hidden (origin 1)'''
  if 'hidden' not in table_info:
    return 1
  for col_no,col_data in enumerate(table_info['columns']):
    if col_data['name'] in table_info['hidden']:
      continue
    return col_no+1

def write_table(wb,target_sheet,table_name,df,groups=None,title_row=None,edit_checks=None,table_map=None):
  '''
  Write the dataframe to the worksheet as a table, formatting numbers
  If groups provided, create folding based on groups.
  If edit_checks are provided, insert the formula into cells to the right on the 1st data row of the table
  Reads the config file to determine some values.

  args:
    wb: the openpyxl workbook to write to
    target_sheet: the name of the worksheet in the workbook
    table_name: the name of the table to write into the target_sheet
    df: The prepared dataframe that has the correct columns
    groups: a list of 3 element lists, each representing a grouping. The elements are level, start, end
    title_row: to allow for stacking of multiple tables on a sheet. If provided it will override the config value, default None
    edit_checks: a list of edit checks which each contain a list of "for_columns" and the formula.
    table_map: a dict of the tables available so far - used by edit_checks

  returns: the workbook

  raises: IndexError if table not in config for the given sheet
  '''
  df.reset_index(drop=True,inplace=True) # rely on the index to figure row on sheet.
  try:
    config=read_config()
  except FileNotFoundError as e:
    raise FileNotFoundError('workbook or config file not found ') from e
  tables=config['sheets'][target_sheet]['tables']
  tx=[table_info['name'] for table_info in tables].index(table_name)
  sg=config['sheets'][target_sheet]['sheet_group']
  table_style=config['sheet_groups'][sg]['table_style']
  try:
    table_info=tables[tx]
  except IndexError as e:
    raise IndexError('Bad table reference: {}'.format(table_name)) from e

  key_values={'title_row':1,'start_col':1,'include_years':False}
  for k in key_values:   #take specified or default items
    if k in table_info:
      key_values[k]=table_info[k]
  if title_row is not None:
    key_values['title_row']=title_row

  title_cell=f'{get_column_letter((key_values["start_col"]-1)+first_not_hidden(table_info))}{key_values["title_row"]}'

  ws=wb[target_sheet]
  ws[title_cell].value=table_info['title']
  ws[title_cell].font=Font(name='Calibri',size=16,color='4472C4')
  table_start_row=key_values['title_row']+1 # the heading of the table (normally excel row 2 )
  table_start_col=key_values['start_col']

  col_defs=pd.DataFrame(table_info['columns']).set_index('name') # column name is index, attributes are columns

  # Place the headings
  for cx,cn in enumerate( df.columns):
    ws.cell(table_start_row,column=table_start_col+cx).value=cn
  # place the values
  fin_format=BUILTIN_FORMATS[41] #''_(* #,##0_);_(* \\(#,##0\\);_(* "-"_);_(@_)'
  first_field=None #first non-formula field - used to determine if its a rate
  for ix,values in df.iterrows(): # the indexes (starting at zero), and the data values
    for cx,cn in enumerate( df.columns):
      if cn in values:
        if first_field is None: # pick this field if its not a formula
          if not is_formula(values[cn]):
            first_field=cn
        rix=ix+table_start_row+1
        cix=table_start_col+cx

        if not is_formula(values[cn]):
          ws.cell(row=rix,column=cix).value=values[cn] # set the value
        else: # its a formula, 
          # so called "future functions" need prefixes as do certain parameters
          formula=prepare_formula(values[cn])
          # flag to force to array formulas to avoid excel adding @ at inconvenient places
          ff=formula != values[cn] 
          # determine if it is an array formula by looking for table column headers with brackets
          regex_column=r'[A-Za-z_]+(\[\[?[ A-Za-z0-9]+\]?\])'
          pattern=re.compile(regex_column)
          matches=pattern.findall(formula) 
          if (len(matches)>0) or ff : # looks like a dynamic formula
            ref=get_column_letter(cix)+str(rix)
            ws[ref] = ArrayFormula(ref,formula) # assume that the formula collapses to a single value
          else:
            ws.cell(row=rix,column=cix).value=formula # set the formula like a regular value  

        # formats for Y columns  
        if cn.startswith('Y')and cn[1:].isnumeric(): 
          # by default use the fin format
          ws.cell(row=rix ,column=cix).number_format=fin_format
          
          # determine if format should be overwritten: for percentage or integer
          # check table title, then line name - last word is key
          fmt_map={'ratios':FORMAT_PERCENTAGE_00,'rate':FORMAT_PERCENTAGE_00,
                    'pct':FORMAT_PERCENTAGE_00,'percent':FORMAT_PERCENTAGE_00,'tax table':FORMAT_NUMBER}
          pat='.*\\b(%s).*'%('|'.join(fmt_map.keys())) # TODO make balance key have a separator and add 2nd \b here
          prog=re.compile(pat)
          line_name=values[first_field]
          if line_name is None: # when no data is provided, there is a row of Nones
            line_name=''
          line_name=line_name.split(':')[-1].strip().lower()
          for legend in table_info['title'].strip().split(' ')[-1].lower(),line_name:
            rx_result=prog.match(legend)
            if rx_result:
              kw=rx_result.group(1)
              special_fmt=fmt_map[kw]
              ws.cell(row=rix,column=cix).number_format=special_fmt
              break

        else: # the non year columns
          if 'horiz' in col_defs.columns:
            horiz=col_defs.loc[cn].horiz
            if pd.notna(horiz):
              ws.cell(row=rix,column=cix).alignment = Alignment(horizontal=horiz)
          if 'number_format' in col_defs.columns: # number formats for non years.
            num_fmt=col_defs.loc[cn].number_format
            if pd.notna(num_fmt):
              if not isinstance(num_fmt,str):
                num_fmt=BUILTIN_FORMATS[num_fmt]
              ws.cell(row=rix,column=cix).number_format=num_fmt



  if groups is not None: # the folding groups
    # set up the row groups highest level to lowest level
    # outline levels are +1 to our levels here
    # 0 here = 1 in Excel - such as Income, Expense
    # 1 here = 2 in Excel - such as I, T and X

    def getlev (e):
      '''utility to aid with setting up groups'''
      return e[0]
    groups.sort(key=getlev)
    for grp in groups:
      # use the start of the group to blank out numeric columns
      for cix,cn in enumerate(df.columns):
        if cn[0]=='Y':
          ws.cell(row=grp[1],column=cix).number_format='###'
      ws.row_dimensions.group(grp[1],grp[2],outline_level=grp[0], hidden=grp[0]>2)


  # making the table
  top_left=f'{get_column_letter(key_values["start_col"])}{table_start_row}'
  bot_right=f'{get_column_letter((key_values["start_col"]-1)+df.shape[1])}{table_start_row+df.shape[0]}'
  rng=f'{top_left}:{bot_right}'
  tab = Table(displayName=table_info['name'], ref=rng)
  # Add a builtin style with striped rows and banded columns
  # The styles are seen on the table tab in excel broken into Light, Medium and Dark.
  # The number seems to be the index in that list (top to bottom, left to right, origin 1)
  tab.tableStyleInfo = TableStyleInfo(name=table_style,  showRowStripes=True)
  ws.add_table(tab)
  logger.info('table {} added to {}'.format(table_info['name'],target_sheet))

   # apply the edit checks, if any
  if edit_checks is not None:
    ws.data_validations.dataValidation.clear() # remove prior items
    base_col=(key_values["start_col"]-1)+df.shape[1]
    for ecx,edit_check in enumerate(edit_checks):
      ws_col=base_col+2*(ecx+1) # get coordinates where the formula will go
      ws_row=table_start_row+1
      ws.cell(row=ws_row-1,column=ws_col).value=f'CHOICES-{ecx}'
      formula=edit_check['formula']
      # starting in openpyxl 3.1 there is array formula support - CSE style only, not dynamic
      # the caller must know how many cells will be contained
      # so evaluate the criteria part of the formula
      table,_,criteria=filter_parser(formula) 
      #... against the data in the source
      source_ws=wb[table_map[table]]
      ref_df=df_for_range(worksheet=source_ws,range_ref=source_ws.tables[table].ref)
      sel=eval_criteria(criteria,ref_df)

      # compute the range of the output of the formula
      col=get_column_letter(ws_col)
      address = f'${col}${ws_row}' 
      ref='%s%d:%s%d'%(col,ws_row,col,ws_row+sel.sum()-1)

      formula=prepare_formula(formula)
      pass
      # place the array formula and its output range in the workbook
      ws[address]=ArrayFormula(ref,formula)
      
      # point to the array for the data validations
      # the trailing # is for excel to know to reference the entire dynamic array
      dv = DataValidation(type="list", formula1=f"_xlfn.ANCHORARRAY({address})", allow_blank=True) 
      for tbl_col in edit_check['for_columns']: # apply to the columns specified
          col_let=get_column_letter((key_values['start_col'])+list(df.columns).index(tbl_col))
          row_refs=[table_start_row+1,table_start_row+df.shape[0]]
          cells='%s%d:%s%d'%(col_let,row_refs[0],col_let,row_refs[1])
          dv.add(cells)
      ws.add_data_validation(dv)


  # add conditional formatting if any
  if  'highlights' in table_info:
    # set range to without the heading
    top_left=f'{get_column_letter(key_values["start_col"])}{table_start_row+1}'
    rng=f'{top_left}:{bot_right}'
    for _,cf in table_info['highlights'].items():
      font=fill=border=None
      # ignore unsupported options
      if 'font' in cf:
        # only support bold, italic and color
        codes={'bold':False,'italic':False, 'color':'000000'}
        for fnt_code, fnt_val in cf['font'].items():
          codes[fnt_code]=fnt_val
        font=Font(bold=codes['bold'],italic=codes['italic'],color=codes['color'])
      if 'fill' in cf:
        codes={'bgcolor':'000000'} # only option
        for fill_code, fill_val in cf['fill'].items():
          codes[fill_code]=fill_val
        fill=PatternFill(start_color=codes['bgcolor'],end_color=codes['bgcolor'])
      if 'border' in cf:
        codes={'edges':None,'style':None,'color': '000000'}
        for bord_code,bord_val in cf['border'].items():
          codes[bord_code]=bord_val
        if codes['edges'] is not None:
          edges=codes['edges']
          if isinstance(edges,str):
            edges=[edges]
          side=Side(border_style=codes['style'],color=codes['color'])
          edge_codes={ 'left':None,'right':None,'top':None,'bottom':None}
          for edge in edges:
            edge_codes[edge]=side
          border=Border(left=edge_codes['left'],right=edge_codes['right'],top=edge_codes['top'],bottom=edge_codes['bottom'])

      dxf= DifferentialStyle(font=font,fill=fill,border=border)
      if 'keys' in cf:
        for key in cf['keys']:
          formula=cf['formula'].format(key)
          rule=Rule(type='expression',dxf=dxf,stopIfTrue=False)
          rule.formula=[formula]
          ws.conditional_formatting.add(rng,rule)
      else:
        rule=Rule(type='expression',dxf=dxf,stopIfTrue=False)
        rule.formula=[cf['formula']]
        ws.conditional_formatting.add(rng,rule)

  return wb

def columns_for_table(wb,sheet,table_name,config):
  '''Get the column definitions for a table based on the config

  args:
    wb: the name of the workbook which may hold the correct first forecast year,
      but if its not there, we get it from the config.
    sheet: the name of the sheet where the table is
    table_name: table name, like 'tbl_x_x'
    config: the config dict

  returns:
    A pandas dataframe with the name of the columns and attributes as columns
    The index can be used with an offset to locate the table on the worksheet.  The first index is zero.

  '''
  ffy=get_f_fcast_year(wb,config)
  for table_info in config['sheets'][sheet]['tables']:
    if table_info['name']!=table_name:
      continue
    years=range(config['start_year'],config['end_year']+1)
    # limit years as needed for actual only
    ao=False
    if 'actual_only' in table_info:
      ao = table_info['actual_only']
    if ao:
      years=[y for y in years if y < ffy]
    years=['Y{}'.format(y) for y in years]
    col_defs=[{'name':'default','width':config['year_column_width']}]
    col_defs+=table_info['columns']
    iy=False
    if 'include_years' in table_info:
      iy=table_info['include_years']
    if iy:
      for y in years:
        col_defs.append({'name':y,'width':config['year_column_width']})
    df=pd.DataFrame(col_defs)
    df=pd.DataFrame.fillna(df,method='ffill') # forward fill missing values
    df.width=df.width.astype(int)
    df=df.drop(0) # remove the default
    df.reset_index(drop=True,inplace=True)
    hide_these=[]
    if 'hidden' in table_info:
      hide_these=table_info['hidden']
    df['hidden']= df.name.isin(hide_these)
    return df

def conform_table(df,columns):
  '''Add any missing table columns and put the columns in the order
  defined by the config in the table_info.
  This also removes any extraneous columns.

  args:
    df: a dataframe which may have missing columns or columns in wrong order
    columns: a pandas series of column names (such as from the output of columns_for_table)

  returns:
    a dataframe with exactly the right set of columns in the right order.
    '''
  sel=~columns.isin(df.columns)
  for col in columns.loc[sel]: # create the missing columns
    df[col]=None
  return df[columns]# possibly alter the order
