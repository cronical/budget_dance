"""Utility to add tabs to the workbook and sets the tab colors

May be modified to add new tabs and re-run.  Does not replace or delete any tabs, only adds them
"""

from openpyxl import Workbook
from openpyxl import load_workbook
from util.logs import get_logger
logger=get_logger(__file__)

target='data/fcast.xlsm'
tabs=['accounts','balances','invest_actl','transfers_actl','iande_actl','hsa','401K','bank_actl','iande_map','iande','aux','capgn','taxes','tax-est','tables','retirement','retireparms','utility']
colors=['4FA3DB','4FA3DB','58BD2D','58BD2D','58BD2D','58BD2D','58BD2D','58BD2D','BDAE2D','BDAE2D','BDAE2D','BD710D','BD710D','BD710D','BD710D','910DBD','910DBD','7E0505']
assert len(colors)== len(tabs)
tc=zip(tabs,colors)
wb=load_workbook(filename = target,keep_vba=True)
sheets=wb.sheetnames
logger.info('{} existing sheets in {}'.format(len(sheets),target))
for t in tc:
  try:
    x=sheets.index(t[0]) # is it already there?
    ws = wb[t[0]]
  except ValueError:
    ws=wb.create_sheet(t[0])
    logger.info('sheet {} added'.format(t[0]))
  ws.sheet_properties.tabColor= t[1]
wb.save(target)
logger.info('workbook {} saved'.format(target))  