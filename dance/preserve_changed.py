#! /usr/bin/env python
'''Look for values with no formulas and save them off for eventual reload'''
import argparse
import json
import re
from math import isnan
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils.cell import coordinate_to_tuple
import pandas as pd
from dance.util.files import read_config
from dance.util.logs import get_logger
from dance.util.tables import df_for_table_name, table_as_df

config=read_config() 
logger=get_logger(__file__)

def is_formula(item):
  if not isinstance(item,str): return False
  return item.startswith('=')

def is_fcst(col,ffy):
  '''Return true if col is a forecast '''
  r=False
  pat=re.compile("(Y[0-9]{4,4})")
  m=pat.match(col)
  if m: # todo allow edit of enumerated fields based on config
    if m.group(1)==col:
      if int(col[1:])>=ffy:
        r=True  
  return r

def save_sparse(config,workbook,path):
  '''write cells that meet criteria out to a json file
  tables are eligible if config has preserve method sparse
  columns are eligible if a forecast year or in preserve non-year-cols
  cell is eligible if it is not a formula'''
  out=[]
  counters={}
  ffy=config['first_forecast_year']
  for _,sheet_info in config['sheets'].items():
    for table_info in sheet_info['tables']:
      if 'preserve' in table_info:
        if table_info['preserve']['method']=='sparse':
          non_year_cols=[]
          if 'non-year-cols' in table_info['preserve']:
            non_year_cols=table_info['preserve']['non-year-cols']
          table=table_info['name']
          counters[table]=0
          df= df_for_table_name(table,workbook,data_only=False)# key is in index
          for ix,row in df.iterrows():
            for col,item in row.items():
              if (is_fcst(col,ffy) or (col in non_year_cols)):
                if pd.notnull(item):
                  if not is_formula(item): # will save ytd forwarded items, but no matter
                    out.append({'table':table,'row':ix,'col':col,'value':item})
                    counters[table]+=1
          logger.info('Found %d items from table %s'%(counters[table],table))
  with open (path,'w',encoding='utf-8') as f:
    json.dump(out,f,ensure_ascii=False,indent=2)
  logger.info('Wrote %d items to %s'%(len(out),args.path))

def load_sparse(config,workbook,path):
  '''copy items from saved json file back into various tables
  Only writes values that differ from what is aready in the sheet and are in forecast periods'''
  comment=Comment('Preserved from previous run of preservation utility','BudgetDance')
  wb=load_workbook(filename = workbook)
  with open (path,encoding='utf-8')as f:
    df_points=pd.read_json(f,orient='records')
  years=df_points.col.str.extract("Y(\d{4})").squeeze()
  sel=years.isnull() # mark items that are not years
  keep=years.loc[~sel].apply(int)>=config['first_forecast_year']
  sel.loc[~sel]=keep
  logger.info(f"Removing {(~sel).sum()} rows that are prior to the first forecast year")
  df_points=df_points.loc[sel]
  counters={}
  for table in pd.unique(df_points.table):
    counters[table]=0
    df_table,ws_name,ref=table_as_df(wb,table)
    ws=wb[ws_name]
    for _,row in df_points.loc[df_points.table==table].iterrows():
      top_left=coordinate_to_tuple(ref.split(':')[0]) # in Excel origin 1
      # in Python origin 0, so adding to top_left is also in Excel origin 1
      rx=top_left[0]+df_table.index.get_loc(row['row']) +1 # +1 to skip the table heading 
      cx=top_left[1]+df_table.columns.get_loc(row['col'])+1 # +1 since we moved the column to the index
      val=row['value']
      curr_val=ws.cell(row=rx,column=cx).value
      if val != curr_val:
        if not (isnan(val) and (curr_val is None)):
          ws.cell(row=rx,column=cx,value=val)
          ws.cell(row=rx,column=cx).comment=comment  
          # the above causes UserWarning: Duplicate name: 'xl/drawings/commentsDrawing1.vml'
          # when this is run after ytd.py but not if its run 1st!!!
          counters[table]+=1
    logger.info('Wrote %d values into table %s'%(counters[table],table))
  wb.save(workbook)

def main():
  defaults={'workbook':config['workbook'],'storage':'./data/preserve.json'}
  parser = argparse.ArgumentParser(description ='Copies changed data from tables to file or from file to various tables')
  group=parser.add_mutually_exclusive_group(required=True)
  group.add_argument('-s','--save',help='saves data from the current tab to the file',action='store_true')
  group.add_argument('-l','--load',help='loads the file data workbook',action='store_true')
  parser.add_argument('-w','--workbook',default=defaults['workbook'],help='Target workbook. Default: '+defaults['workbook'])
  parser.add_argument('-p','--path',default=defaults['storage'] ,help='The path and name of the storage file. Default='+defaults['storage'])
  args=parser.parse_args()
    
  if args.save:
    save_sparse(config,args.workbook,args.path)

  if args.load :
    load_sparse(config,args.workbook,args.path)

if __name__ == '__main__':
  main()