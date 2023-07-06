#! /usr/bin/env python
'''Look for values with no formulas and save them off for eventual reload'''
import argparse
import json
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils.cell import coordinate_to_tuple
import pandas as pd
from dance.util.files import read_config
from dance.util.logs import get_logger
from dance.util.tables import df_for_table_name, table_as_df

logger=get_logger(__file__)
comment=Comment('Recovered from previous run of preservation utility',author='system')

def is_formula(item):
  if not isinstance(item,str): return False
  return item.startswith('=')
def save_sparse(config,workbook,path):
  '''write cells that meet criteria out to a json file'''
  out=[]
  counters={}
  ffy=config['first_forecast_year']
  for table in ['tbl_iande','tbl_balances']:# TODO list tables
    counters[table]=0
    df= df_for_table_name(table,workbook,data_only=False)# key is in index
    for ix,row in df.iterrows():
      for col,item in row.items():        
        if col.startswith('Y'): # todo allow edit of enumerated fields based on config
          if int(col[1:])>=ffy:
            if not is_formula(item): # will save ytd forwarded items, but no matter
              if item is not None:
                out.append({'table':table,'row':ix,'col':col,'value':item})
                counters[table]+=1
    logger.info('Found %d items from table %s'%(counters[table],table))
  with open (path,'w',encoding='utf-8') as f:
    json.dump(out,f,ensure_ascii=False,indent=2)
  logger.info('Wrote %d items to %s'%(len(out),args.path))

def load_sparse(config,workbook,path):
  '''copy items from saved json file back into various tables'''
  wb=load_workbook(filename = workbook,keep_vba=True)
  with open (path,encoding='utf-8')as f:
    df_points=pd.read_json(f,orient='records')
  counters={}
  for table in pd.unique(df_points.table):
    counters[table]=0
    df_table,ws_name,ref=table_as_df(wb,table)
    ws=wb[ws_name]
    for _,row in df_points.loc[df_points.table==table].iterrows():
      top_left=coordinate_to_tuple(ref.split(':')[0]) # in Excel origin 1
      # in Python origin 0, so adding to top_left is also in Excel origin 1
      x=top_left[0]+df_table.index.get_loc(row['row']) +1 # +1 to skip the table heading 
      y=top_left[1]+df_table.columns.get_loc(row['col'])+1 # +1 since we moved the column to the index
      val=row['value']
      ws.cell(row=x,column=y,value=val)
      counters[table]+=1
    logger.info('Wrote %d values into table %s'%(counters[table],table))
  wb.save(args.workbook)

if __name__ == '__main__':
  defaults={'workbook':'data/test_wb.xlsm','storage':'./data/preserve.json'}# TODO fcast
  parser = argparse.ArgumentParser(description ='Copies changed data from tables to file or from file to various tables')
  group=parser.add_mutually_exclusive_group(required=True)
  group.add_argument('-s','--save',help='saves data from the current tab to the file',action='store_true')
  group.add_argument('-l','--load',help='loads the file data workbook',action='store_true')
  parser.add_argument('-w','--workbook',default=defaults['workbook'],help='Target workbook. Default='+defaults['workbook'])
  parser.add_argument('-p','--path',default=defaults['storage'] ,help='The path and name of the storage file. Default='+defaults['storage'])
  args=parser.parse_args()
  config=read_config()

  
  if args.save:
    save_sparse(config,args.workbook,args.path)

  if args.load :
    load_sparse(config,args.workbook,args.path)
