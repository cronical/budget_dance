#! /usr/bin/env python
'''Save or load YTD values'''
import argparse
import json
import sys

from openpyxl import load_workbook,Workbook
from openpyxl.comments import Comment
import pandas as pd
from dance.util.files import read_config
from dance.util.logs import get_logger
from dance.util.tables import df_for_table_name

config=read_config() 
logger=get_logger(__file__)

comment=Comment('Estimated based on year to date data.','ytd.py')
defaults={'workbook':config['workbook'],'storage':'./data/ytd_data.json'}

def main():
  parser = argparse.ArgumentParser(description ='Copies data from ytd tab to file or from file to ytd tab and iande.')
  parser.add_argument('-s','--save',help='saves data from the current tab to the file',action='store_true')
  parser.add_argument('-l','--load',help='loads the file data to the current tab',action='store_true')
  parser.add_argument('-f','--forward',help='carries the projected values to the first forecast year in the iande table'
                      ,action='store_true')
  parser.add_argument('-w','--workbook',default=defaults['workbook'],help='Target workbook. Default='+defaults['workbook'])
  parser.add_argument('-p','--path',default=defaults['storage'] ,help='The path and name of the storage file. Default='+defaults['storage'])
  args=parser.parse_args()
  tgt_year='Y%04d'%config['first_forecast_year']

  if args.save:
    save_ytd(args)

  wb=load_workbook(filename = args.workbook)
  counters={}
  if args.load or args.forward:
    j_df=read_json_file(args.path,tgt_year=tgt_year)
    # now do load and forward if called for
    if(args.load):
      wb,counters=load_ytd(j_df,wb,counters)
    if(args.forward):
      wb,counters=forward_ytd(j_df,wb,counters,tgt_year)      
    if sum(counters.values())>0:
      wb.save(args.workbook)

def read_json_file(path:str,tgt_year:str):
  with open (path,encoding='utf-8')as f:
    j_df=pd.read_json(f,orient='index')
    if not j_df.columns[0].startswith(tgt_year):
      logger.warning(f"Skipping YTD load since year is {tgt_year} and year to date data is as of {j_df.columns[0]}")
      sys.exit(-1)
    return j_df

def load_and_forward(config:dict, path:str):
  """For non-command line use, perform load and foward
  """
  tgt_year='Y%04d'%config['first_forecast_year']
  j_df=read_json_file(path,tgt_year)
  wb=load_workbook(filename = config['workbook'])
  counters={}
  wb,counters=load_ytd(j_df=j_df,wb=wb,counters=counters)
  wb,counters=forward_ytd(j_df=j_df,wb=wb,counters=counters,tgt_year='Y%04d'%config['first_forecast_year'])
  if sum(counters.values())>0:
    wb.save(defaults['workbook'])

def load_ytd(j_df,wb:Workbook,counters:dict):
  """load saved values into current table
  """
  parms={'tab':'current','sources':['Factor','Add'],'targets':['Factor','Add']}
  return place_values(j_df,wb,counters,parms)

def forward_ytd(j_df,wb:Workbook,counters:dict,tgt_year:str):
  """forward saved values into iande table for the target year
  """
  parms={'tab':'iande','sources':['Year'],'targets':[tgt_year]}
  return place_values(j_df,wb,counters,parms)

def save_ytd(args):
    """save the values from the current table into the json file
    """
    curr_df= df_for_table_name('tbl_current',args.workbook,data_only=True)# key is in index
    ytd_fld_ix=list(curr_df.columns.str.startswith('Y')).index(True)
    ytd_fld=curr_df.columns[ytd_fld_ix]
    key_val=curr_df.loc[(curr_df.is_leaf==1) & (curr_df.Factor.notna() | curr_df.Add.notna()),[ytd_fld,'Factor','Add','Year']]
    key_val=key_val.fillna(0) 
    with open (args.path,'w',encoding='utf-8') as f:
      data=json.loads(key_val.to_json(orient='index'))
      json.dump(data,f,ensure_ascii=False,indent=2)
    logger.info('Wrote %d items to %s'%(len(key_val),args.path))

def place_values(j_df, wb,  counters, control):
  """ Place the values into either the current tab or the iande tab
  returns modified workbook, counters
  """
  row_offset=3 # 1 to change origin, 1 for title, 1 for heading
  col_offset=2 # 1 to change origin, 1 since key is in the index not a field
  src_tgt=list(zip(control['sources'],control['targets']))
  
  ws_name=control['tab']
  table='tbl_'+ws_name
  counters[table]=0
  ws=wb[ws_name]
  tgt_df=df_for_table_name(table,wb,data_only=True)
  idx = tgt_df.index.to_list() # key is in index
  for ix,values in j_df.iterrows():
    for src,tgt in src_tgt:
      cx=col_offset+tgt_df.columns.to_list().index(tgt)
      rx=row_offset+idx.index(ix)
      val=values[src]
      if ws_name != 'iande':
        ws.cell(row=rx,column=cx).value=val
        counters[table]+=1
      else: # for iande if val is not already there, set it and comment
        curr_val=ws.cell(row=rx,column=cx).value
        if val != curr_val:
          ws.cell(row=rx,column=cx,value=val)
          ws.cell(row=rx,column=cx).comment=comment
          counters[table]+=1
    if ws_name=='current':
      # recompute reprojected year, in case the excel calc has not yet run.
      yx=col_offset+list(tgt_df.columns.str.startswith('Y')).index(True) # locate the ytd date column in the latest tsv
      j_df.at[ix,'Year']=ws.cell(row=rx,column=yx).value*values['Factor']+values['Add'] 
  logger.info('Wrote %d values into table %s'%(counters[table],table))
  return wb,counters
      
if __name__ == '__main__':
  main()