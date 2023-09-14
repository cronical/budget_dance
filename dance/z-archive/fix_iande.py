"""This is clear out some calculations that are extened through forecast time referencing iande-actl.

A one time fix to a programming error, but it could be useful again, so leaving it around
"""


from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
import openpyxl.utils.cell as ut
from util.logs import get_logger
from dance.util.files import read_config
logger=get_logger(__file__)
source=read_config()['workbook']
target = 'fcast2.xlsm'
wb = load_workbook(filename = source, read_only=False)
logger.info('loaded workbook from {}'.format(source))
sheet_name= 'iande'
new_sheet_name='iande_2'
tab_src='tbl_iande'
tab_tgt='tbl_iande_2'
all_sheets=wb.sheetnames
ix= wb.sheetnames.index(sheet_name)
src=wb[sheet_name]
if new_sheet_name in wb.sheetnames:  wb.remove(new_sheet_name)
tgt=wb.create_sheet(new_sheet_name,ix)
tgt.sheet_properties.tabColor=src.sheet_properties.tabColor
logger.info('created worksheet {}'.format(new_sheet_name))


tab=src.tables[tab_src]
assert tab_src == tab.name
rb=ut.range_boundaries(tab.ref)
ix=tab.column_names.index('Y2022')

for r1, r2 in zip(src.iter_rows(min_row=1, min_col=1, max_col=rb[2], max_row=rb[3]) \
  , tgt.iter_rows(min_row=1, min_col=1, max_col=rb[2], max_row=rb[3])):
  for c1, c2, cix in zip(r1, r2, range(rb[2])):
    val=c1.value
    if cix >= ix:
      if type(val) is str:
        if -1 != val.find('tbl_iande_actl'):
          val=None
    c2.value = val

wb.save(target)
logger.info('saved to {}'.format(target))