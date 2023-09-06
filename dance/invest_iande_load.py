#! /usr/bin/env python
'''
Prepare the data from the Investment IandE report from which is in the file invest-iande.tsv.

'''
import argparse
from openpyxl import load_workbook

import pandas as pd

from dance.util.books import fresh_sheet, col_attrs_for_sheet,set_col_attrs
from dance.util.files import tsv_to_df, read_config
from dance.util.sheet import df_for_range
from dance.util.tables import  columns_for_table, write_table,get_f_fcast_year,conform_table
from dance.util.logs import get_logger
from dance.util.xl_formulas import actual_formulas,forecast_formulas,dyno_fields,this_row

logger=get_logger(__file__)

def y_years(col_name):
  '''Given a column name return it unless its a number, in that case prepend "Y".'''
  if isinstance(col_name,str):
    return col_name
  return f'Y{col_name}'

def read_and_prepare_invest_iande(workbook,data_info,f_fcast=None):
  '''  Read investment income and expense actual data from file into a dataframe

  args:
    workbook: name of the workbook (where to get the accounts data and the unrealized values used to calc market gn rates)
    data_info: dict that has a value for path used to locate the input file, which contains transaction data for a series of years
    f_fcast: Optional. The first forecast year as Ynnnn. If none, will read from the workbook file. Default None.
    table_map: the dict that maps tables to worksheets. Required for the initial setup as it is not yet stored in file

  returns:  A summarized dataframe which has years for columns, where summed values are.
    It also has columns for the Investment account name and the I and E category.

  raises:
    FileNotFoundError: if the input file does not exist.
  '''
    # get the workbook from file
  try:
    wb = load_workbook(filename = workbook, read_only=False, keep_vba=True)
    logger.info('loaded workbook from {}'.format(workbook))
    config=read_config()
  except FileNotFoundError:
    raise FileNotFoundError(f'file not found {workbook}') from None
  if f_fcast is None:
    f_fcast='Y%d' % get_f_fcast_year(wb,config) # get the first forecast year from the general state table or config as available
  logger.info ('First forecast year is: %s',f_fcast)

  input_file=data_info['path']
  try:
    df=tsv_to_df (input_file,nan_is_zero=False,skiprows=3,string_fields=['Check#','Description','Category','Tags','C'])
    logger.info('loaded dataframe from {}'.format(input_file))
  except FileNotFoundError as e:
    raise f'file not found {input_file}' from e
  df=df.loc[df.Account !='Total'] # remove total line
  df=df.dropna(how='all')# and blank row after total
  df=df[['Account','Date','Category','Amount']]
  df['Year']=df.Date.dt.year
  vals_df=pd.pivot_table(df,index=['Account','Category'],aggfunc='sum',values='Amount',columns='Year')
  vals_df.reset_index(inplace=True) # Account and category are now columns

  # add rows for unrealized gains - actuals will be set by formulas
  # pull all the accounts not just the ones that are in the performance report
  ws=wb['accounts']
  ref=ws.tables['tbl_accounts'].ref
  accounts_df=df_for_range(ws,ref)
  ia_df=vals_df.loc[vals_df.index <0] # zero rows, same cols
  ia_df=ia_df.append(pd.DataFrame(accounts_df.index[accounts_df.Type=='I'],columns=['Account']))
  ia_df['Category']="Unrlz Gn/Ls"
  vals_df=pd.concat([vals_df,ia_df])
  vals_df.reset_index(drop=True,inplace=True)

  full_categories=list(vals_df.Category.str.split(':'))
  vals_df['Category']=[':'.join(a[-2:])for a in full_categories]  # last 2 parts of the category are used
  vals_df.fillna(0,inplace=True) # no missing values - use zeros
  vals_df.insert(loc=2,column='IorE',value=[a[0]for a in full_categories],allow_duplicates=True)# Mark and I, X or T
  sel=(vals_df.Category=='Unrlz Gn/Ls')
  vals_df.loc[sel,'IorE']='U' # so it doesn't get included in realized gains
  vals_df.insert(0,column='Key',value=vals_df.Account+':'+vals_df.Category) # full key
  vals_df.rename(columns=y_years,inplace=True) # use the y_years format for column names
  for y in range(int(f_fcast[1:]),config['end_year']+1): # add columns for forecast years
    vals_df['Y{}'.format(y)]=None
  col_def=columns_for_table(wb,'invest_iande_work','tbl_invest_iande_values',config)
  vals_df=conform_table(vals_df,col_def['name'])
  return vals_df
  
def prepare_invest_iande_ratios(df):
  '''Zero out the values in the values dataframe to create the ratios frame'''
  ys=[c for c in df.columns if c.startswith('Y') ]
  df[ys]=0 # ratios to be completed by setup.yaml  
  return df


if __name__ == '__main__':
  parser = argparse.ArgumentParser(description ='Prepares income and expense data to insert into  "tbl_invest_iande_work". ')
  parser.add_argument('--workbook','-w',default='data/test_wb.xlsm',help='Target workbook') # TODO change to fcast.xlsm
  parser.add_argument('--path','-p',default= 'data/invest-iande.tsv',help='The path and name of the input file')

  args=parser.parse_args()
  config=read_config()
  ffy=config['first_forecast_year']
  sheet='invest_iande_work'
  wkb = load_workbook(filename = args.workbook, read_only=False, keep_vba=True)
  wkb=fresh_sheet(wkb,sheet)
  title_row=1
  for table_info in config['sheets'][sheet]['tables']:
    table=table_info['name']
    data_info=table_info['data']
    data_info['path']=args.path
    data=read_and_prepare_invest_iande(workbook=args.workbook,data_info=data_info)
    if table=='tbl_invest_iande_ratios':
      data=prepare_invest_iande_ratios(data)
    data=dyno_fields(table_info,data)
    data=forecast_formulas(table_info,data,ffy) # insert forecast formulas per config
    data=actual_formulas(table_info,data,ffy) # insert actual formulas per config
    wkb= write_table(wkb,target_sheet=sheet,df=data,table_name=table,title_row=title_row)
    title_row=title_row+3+data.shape[0]
  attrs=col_attrs_for_sheet(wkb,sheet,read_config())
  wkb=set_col_attrs(wkb,sheet,attrs)
  wkb.save(args.workbook)

  pass
