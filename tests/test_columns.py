'''tests for the column handling in the tables and books utility modules'''
from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from dance.util.files import read_config
from dance.util.tables import columns_for_table, get_value_for_key,get_f_fcast_year
from dance.util.books import col_attrs_for_sheet

import pytest

@pytest.fixture
def bad_wb():
  return Workbook() # won't have the utility sheet

@pytest.fixture
def good_wb():
  wb=Workbook()
  ws=wb.create_sheet('utility')
  ws.append(['Table','Worksheet'])
  ws.append(['tbl_gen_state','tables'])
  tab=Table(displayName='tbl_table_map',ref='A1:B2')
  ws.add_table(tab)

  ws=wb.create_sheet('tables')
  ws.append(['Item','Value'])
  ws.append(['first_forecast','Y2042'])
  tab=Table(displayName='tbl_gen_state',ref='A1:B2')
  ws.add_table(tab)
  return wb

@pytest.fixture
def fake_config():
  config=read_config()
  config['first_forecast_year']=2037
  config['start_year'] = 2036
  config['end_year'] = 2038
  return config

@pytest.fixture
def std_config():
  return read_config()

@pytest.fixture
def sheet():
  return 'iande_actl'
@pytest.fixture
def table():    
  return 'tbl_iande_actl'

def test_get_value_for_key_wb_invalid(bad_wb):
  with pytest.raises(ValueError):
    _=get_value_for_key(bad_wb,'foobar')

def test_get_value_for_key_wb_ok_bad_key(good_wb):
  with pytest.raises(ValueError):
    _=get_value_for_key(good_wb,'foobar')

def test_get_f_fcast_year_wb_invalid(bad_wb,fake_config):
  ffy=get_f_fcast_year(bad_wb,fake_config)
  assert ffy == fake_config['first_forecast_year']

def test_get_f_fcast_year_wb_ok(good_wb,fake_config):
  ffy=get_f_fcast_year(good_wb,fake_config)
  assert ffy == 2042

def test_columns_for_table_ao(bad_wb,sheet,table,fake_config):
  df=columns_for_table(bad_wb,sheet,table,fake_config)
  assert all([a in df.columns for a in ['name','width','hidden']])
  assert list(df.name)==['Key','Account','Y2036']
  assert sum(df.hidden)==1

def test_columns_for_table_all(bad_wb,fake_config):
  df=columns_for_table(bad_wb,'iande','tbl_iande',fake_config)
  assert all([a in df.columns for a in ['name','width','hidden']])
  assert list(df.name)==['Key','Account','Y2036','Y2037','Y2038']
  assert sum(df.hidden)==1



def test_one_table(bad_wb,std_config):
  attrs=col_attrs_for_sheet(bad_wb,'utility',std_config)
  assert attrs=={1:{'width':20, 'hidden': False},2:{'width':20, 'hidden': False}}

def test_multi_table_flat(bad_wb,std_config):
  attrs=col_attrs_for_sheet(bad_wb,'tax_tables',std_config)
  expected={1:{'width':8, 'hidden': False},2:{'width':14, 'hidden': False},3:{'width':14, 'hidden': False},4:{'width':14, 'hidden': False},
  6:{'width':14, 'hidden': False},7:{'width':18, 'hidden': False},9:{'width':14, 'hidden': False},10:{'width':18, 'hidden': False}}
  assert attrs==expected

def test_multi_table_overlap(bad_wb,std_config):
  attrs=col_attrs_for_sheet(bad_wb,'gen_tables',std_config)
  expected={1:{'width':18, 'hidden': False},2:{'width':18, 'hidden': False},3:{'width':18, 'hidden': False},4:{'width':18, 'hidden': False},
  5:{'width':18, 'hidden': False},6:{'width':18, 'hidden': False},7:{'width':18, 'hidden': False},8:{'width':18, 'hidden': False},
  9:{'width':18, 'hidden': False},10:{'width':18, 'hidden': False},11:{'width':40, 'hidden': False},12:{'width':18, 'hidden': False}}
  assert attrs==expected

def test_all_years(bad_wb,std_config):
  attrs=col_attrs_for_sheet(bad_wb,'iande',std_config)
  expected={1:{'width':70,'hidden':True},2:{'width':30,'hidden':False}}
  for x in range(3, 16):
    expected[x]={'width':12,'hidden':False}
  assert attrs==expected

def test_actl_years(bad_wb,std_config):
  ffy=std_config['first_forecast_year']
  s=std_config['start_year']
  e=std_config['end_year']
  years=[y for y in range(s,e+1) if y <ffy]

  attrs=col_attrs_for_sheet(bad_wb,'iande_actl', std_config)
  expected={1:{'width':35,'hidden':True},2:{'width':35,'hidden':False}}
  base=1+len(expected)
  for x,_ in enumerate(years):
    expected[base+x]={'width':12,'hidden':False}
  assert attrs==expected
